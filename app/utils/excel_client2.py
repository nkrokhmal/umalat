import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
import json
from .openpyxl_wrapper import ExcelBlock


class Cell:
    def __init__(self, row, column, name):
        self.row = row
        self.column = column
        self.name = name


SHEETS = {
    0: 'файл остатки',
    1: 'планирование суточное'
}

COLOURS = {
    'Для пиццы': 'E5B7B6',
    'Моцарелла': 'DAE5F1',
    'Фиор Ди Латте': 'CBC0D9',
    'Чильеджина': 'E5DFEC',
    'Качокавалло': 'F1DADA',
    'Сулугуни': 'F1DADA',
    'Default': 'D9DDDC'
}

ORDER = ['Фиор Ди Латте', 'Чильеджина', 'Моцарелла', 'Сулугуни', 'Для пиццы', 'Качокавалло']

CELLS = {
    'Boiling': Cell(1, 1, 'Тип варки'),
    'FormFactor': Cell(1, 2, 'Форм фактор'),
    'Brand': Cell(1, 3, 'Бренд'),
    'SKU': Cell(1, 4, 'Номенклатура'),
    'FactRemains': Cell(1, 5, 'Факт.остатки, заявка'),
    'NormativeRemains': Cell(1, 6, 'Нормативные остатки'),
    'ProductionPlan': Cell(1, 7, 'План производства'),
    'Volume': Cell(1, 9, 'Объем варки'),
    'Estimation': Cell(1, 10, 'Расчет'),
    'Plan': Cell(1, 11, 'План'),
    'BoilingVolumes': Cell(1, 12, 'Объемы варок'),
    'Name1': Cell(1, 15, 'Фактические остатки на складах - Заявлено, кг:'),
    'Name2': Cell(2, 15, 'Нормативные остатки, кг')
}


COLUMNS = {
    'BoilingVolume': 10,
    'SKUS_ID': 18,
    'BOILING_ID': 19
}

BOILING_LIMITS = {
    'MIN': 6000,
    'MAX': 8000
}


def generate_title(sheet):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(COLUMNS['SKUS_ID'])].hidden = True
    sheet.column_dimensions[openpyxl.utils.get_column_letter(COLUMNS['BOILING_ID'])].hidden = True
    for key, val in CELLS.items():
        sheet.cell(val.row, val.column).value = val.name
    sheet.freeze_panes = sheet['A2']

    sheet.column_dimensions[get_column_letter(CELLS['Brand'].column)].width = 3 * 5
    sheet.column_dimensions[get_column_letter(CELLS['SKU'].column)].width = 10 * 5
    sheet.column_dimensions[get_column_letter(CELLS['Boiling'].column)].width = 5 * 5


def build_plan2(date, df, request_list, plan_path=None):
    filename = '{}_{}.xlsx'.format('plan', date.strftime('%Y-%m-%d'))
    if plan_path is None:
        path = '{}/{}'.format('app/data/plan', filename)
    else:
        path = plan_path
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, sheet_name=SHEETS[0])
        writer.save()

    wb = openpyxl.load_workbook(filename=path)
    sheet_plan = wb.create_sheet(SHEETS[1])
    generate_title(sheet=sheet_plan)

    cur_row, space_rows = 2, 2
    for group_skus in request_list:
        beg_row = cur_row
        block = ExcelBlock(sheet=sheet_plan)
        group_formula = []

        for form_factor in ORDER:
            block_skus = [x for x in group_skus['GroupSKU'] if x['SKU'].form_factor.name == form_factor]
            beg_ff_row = cur_row
            for sku in block_skus:
                formula_plan = "=INDEX('{0}'!$A$5:$DK$265,MATCH($O$1,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$DK$5,0))".format(
                    SHEETS[0], block.sheet.cell(cur_row, CELLS['SKU'].column).coordinate)
                formula_remains = "=INDEX('{0}'!$A$5:$DK$265,MATCH($O$2,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$DK$5,0))".format(
                    SHEETS[0], block.sheet.cell(cur_row, CELLS['SKU'].column).coordinate)

                block.colour = COLOURS[sku['SKU'].form_factor.name]
                block.cell_value(row=cur_row, col=CELLS['Brand'].column, value=sku["SKU"].brand_name)
                block.cell_value(row=cur_row, col=CELLS['SKU'].column, value=sku["SKU"].name)
                block.cell_value(row=cur_row, col=CELLS['FactRemains'].column, value=formula_plan)
                block.cell_value(row=cur_row, col=CELLS['NormativeRemains'].column, value=formula_remains)
                block.cell_value(row=cur_row, col=CELLS['ProductionPlan'].column,
                                 value='=MIN({}, 0)'.format(
                                     block.sheet.cell(cur_row, CELLS['FactRemains'].column).coordinate))

                group_formula.append(block.sheet.cell(cur_row, CELLS['ProductionPlan'].column).coordinate)
                cur_row += 1
            end_ff_row = cur_row - 1
            if beg_ff_row <= end_ff_row:
                block.merge_cells(
                    beg_row=beg_ff_row,
                    end_row=end_ff_row,
                    beg_col=CELLS['FormFactor'].column,
                    end_col=CELLS['FormFactor'].column,
                    value=form_factor,
                    alignment=Alignment(horizontal='center', vertical='center')
                )
        end_row = cur_row - 1
        block.colour = COLOURS['Default']
        block.merge_cells(
            beg_row=beg_row,
            end_row=end_row,
            beg_col=CELLS['Boiling'].column,
            end_col=CELLS['Boiling'].column,
            value='{}% варка, {}{}'.format(
                group_skus["GroupSKU"][0]["SKU"].boilings[0].percent,
                group_skus["GroupSKU"][0]["SKU"].boilings[0].ferment,
                ', без лактозы' if not group_skus["GroupSKU"][0]["SKU"].boilings[0].is_lactose else ''
            ),
            alignment=Alignment(horizontal='center', vertical='center')

        )
        block.cell_value(row=beg_row, col=CELLS['Plan'].column, value='')
        block.cell_value(row=beg_row, col=CELLS['Volume'].column, value=group_skus['Volume'])
        formula_boiling_count = '{}'.format(str(group_formula).strip('[]').replace(',', ' +').replace('\'', "").upper())
        block.cell_value(row=beg_row,
                         col=CELLS['Estimation'].column,
                         value='=-({}) / {}'.format(
                            formula_boiling_count,
                            block.sheet.cell(beg_row, CELLS['Volume'].column).coordinate.upper()))

        block.cell_value(
            row=beg_row,
            col=COLUMNS['SKUS_ID'],
            value=str([x["SKU"].id for x in group_skus["GroupSKU"]])
        )
        block.cell_value(
            row=beg_row,
            col=COLUMNS['BOILING_ID'],
            value=group_skus['BoilingId']
        )
        cur_row += space_rows
    wb.save(path)
    return '{}/{}'.format('data/plan', filename)