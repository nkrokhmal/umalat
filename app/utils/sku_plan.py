import pandas as pd

from app.imports.runtime import *

from openpyxl.styles import Alignment

from app.models import MozzarellaSKU
from app.utils.features.openpyxl_wrapper import ExcelBlock


class Cell:
    def __init__(self, row, column, name):
        self.row = row
        self.column = column
        self.name = name


CELLS = {
    "Boiling": Cell(1, 1, "Тип варки"),
    "FormFactor": Cell(1, 2, "Форм фактор"),
    "Brand": Cell(1, 3, "Бренд"),
    "SKU": Cell(1, 4, "Номенклатура"),
    "FactRemains": Cell(1, 5, "Факт.остатки, заявка"),
    "NormativeRemains": Cell(1, 6, "Нормативные остатки"),
    "NotCalculatedRemainings": Cell(1, 7, "Не учтенные остатки"),
    "ProductionPlan": Cell(1, 8, "План производства"),
    "ExtraPacking": Cell(1, 9, "Дополнительная фасовка"),
    "Volume": Cell(1, 11, "Объем варки"),
    "Estimation": Cell(1, 12, "Расчет"),
    "Plan": Cell(1, 13, "План"),
    "BoilingVolumes": Cell(1, 14, "Объемы варок"),
    "Name1": Cell(1, 16, "Фактические остатки на складах - Заявлено, кг:"),
    "Name2": Cell(2, 16, "Нормативные остатки, кг"),
}

COLUMNS = {"BoilingVolume": 10, "SKUS_ID": 18, "BOILING_ID": 19}


class SkuPlanClient:
    def __init__(self, date, remainings, skus_grouped, template_path, yesterday_boiling_plan_df=pd.DataFrame()):
        self.date = date.strftime("%Y-%m-%d")
        self.filename = "{} План по SKU.xlsx".format(date.strftime("%Y-%m-%d"))
        self.filepath = os.path.join(
            flask.current_app.config["SKU_PLAN_FOLDER"], self.filename
        )
        self.skus_grouped = skus_grouped
        self.remainings = remainings
        self.yesterday_boiling_plan_df = yesterday_boiling_plan_df
        self.space_rows = 2
        self.template_path = template_path
        self.wb = self.create_file()

    def create_file(self):
        shutil.copyfile(self.template_path, self.filepath)
        return openpyxl.load_workbook(self.filepath)

    def fill_remainigs_list(self):
        with pd.ExcelWriter(self.filepath, engine="openpyxl") as writer:
            writer.book = self.wb
            writer.sheets = dict((ws.title, ws) for ws in self.wb.worksheets)
            self.remainings.to_excel(
                writer, sheet_name=flask.current_app.config["SHEET_NAMES"]["remainings"]
            )
            writer.save()
        self.wb = openpyxl.load_workbook(self.filepath)

    def _not_calc_remainings(self, sku_name):
        if self.yesterday_boiling_plan_df.empty:
            return 0
        if sku_name not in self.yesterday_boiling_plan_df.index:
            return 0
        else:
            return self.yesterday_boiling_plan_df.loc[sku_name][0]

    def _set_production_plan(self, obj, value):
        if isinstance(obj, MozzarellaSKU):
            return value if obj.production_by_request else 0
        return value

    def _set_extra_packing(self, obj, value):
        if isinstance(obj, MozzarellaSKU):
            return 0 if obj.packing_by_request else value
        return 0

    def fill_skus(
        self, sku_grouped, excel_client, cur_row, space_factor, sorted_by_weight=False, order=None,
    ):
        beg_row = cur_row
        formula_group = []
        order = order if order is not None else flask.current_app.config["ORDER"]
        for group_name in order:
            block_skus = [x for x in sku_grouped.skus if x.sku.group.name == group_name]
            if block_skus:
                if sorted_by_weight:
                    block_skus = sorted(
                        block_skus, key=lambda k: k.sku.form_factor.relative_weight
                    )

                beg_block_row = cur_row
                for block_sku in block_skus:
                    formula_plan = "=IFERROR(INDEX('{0}'!$A$5:$FG$265,MATCH($P$1,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$FG$5,0)), 0)".format(
                        flask.current_app.config["SHEET_NAMES"]["remainings"],
                        excel_client.sheet.cell(
                            cur_row, CELLS["SKU"].column
                        ).coordinate,
                    )
                    formula_remains = "=IFERROR(INDEX('{0}'!$A$5:$FG$265,MATCH($P$2,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$FG$5,0)), 0)".format(
                        flask.current_app.config["SHEET_NAMES"]["remainings"],
                        excel_client.sheet.cell(
                            cur_row, CELLS["SKU"].column
                        ).coordinate,
                    )

                    excel_client.colour = block_sku.sku.colour[1:]
                    excel_client.draw_cell(
                        row=cur_row,
                        col=CELLS["Brand"].column,
                        value=block_sku.sku.brand_name,
                    )
                    excel_client.draw_cell(
                        row=cur_row,
                        col=CELLS["NotCalculatedRemainings"].column,
                        value=-self._not_calc_remainings(block_sku.sku.name)
                    )

                    excel_client.draw_cell(
                        row=cur_row, col=CELLS["SKU"].column, value=block_sku.sku.name
                    )
                    excel_client.draw_cell(
                        row=cur_row, col=CELLS["FactRemains"].column, value=formula_plan
                    )
                    excel_client.draw_cell(
                        row=cur_row,
                        col=CELLS["NormativeRemains"].column,
                        value=formula_remains,
                    )
                    excel_client.draw_cell(
                        row=cur_row,
                        col=CELLS["ProductionPlan"].column,
                        value=self._set_production_plan(
                            block_sku.sku,
                            "=MIN({} - {}, 0)".format(
                                excel_client.sheet.cell(
                                    cur_row, CELLS["FactRemains"].column
                                ).coordinate,
                                excel_client.sheet.cell(
                                    cur_row, CELLS["NotCalculatedRemainings"].column
                                ).coordinate,
                            ),
                        ),
                    )
                    excel_client.draw_cell(
                        row=cur_row,
                        col=CELLS["ExtraPacking"].column,
                        value=self._set_extra_packing(
                            block_sku.sku,
                            "=MIN({}, 0)".format(
                                excel_client.sheet.cell(
                                    cur_row, CELLS["FactRemains"].column
                                ).coordinate
                            ),
                        ),
                    )

                    formula_group.append(
                        excel_client.sheet.cell(
                            cur_row, CELLS["ProductionPlan"].column
                        ).coordinate
                    )
                    cur_row += 1
                end_block_row = cur_row - 1
                if beg_block_row <= end_block_row:
                    excel_client.merge_cells(
                        beg_row=beg_block_row,
                        end_row=end_block_row,
                        beg_col=CELLS["FormFactor"].column,
                        end_col=CELLS["FormFactor"].column,
                        value=group_name,
                        alignment=Alignment(
                            horizontal="center", vertical="center", wrapText=True
                        ),
                    )
        end_row = cur_row - 1
        excel_client.colour = flask.current_app.config["COLORS"]["DefaultGray"][1:]
        excel_client.merge_cells(
            beg_row=beg_row,
            end_row=end_row,
            beg_col=CELLS["Boiling"].column,
            end_col=CELLS["Boiling"].column,
            value=sku_grouped.boiling.to_str(),
            alignment=Alignment(horizontal="center", vertical="center", wrapText=True),
        )
        excel_client.draw_cell(
            row=beg_row, col=CELLS["Volume"].column, value=sku_grouped.volume
        )
        formula_boiling_count = "{}".format(
            str(formula_group).strip("[]").replace(",", " +").replace("'", "").upper()
        )

        excel_client.draw_cell(
            row=beg_row,
            col=CELLS["Estimation"].column,
            value="=-({}) / {}".format(
                formula_boiling_count,
                excel_client.sheet.cell(
                    beg_row, CELLS["Volume"].column
                ).coordinate.upper(),
            ),
        )

        excel_client.draw_cell(
            row=beg_row,
            col=CELLS["Plan"].column,
            value="=ROUND({}, 0)".format(
                excel_client.sheet.cell(
                    beg_row, CELLS["Estimation"].column
                ).coordinate.upper()
            ),
        )

        excel_client.draw_cell(
            row=beg_row,
            col=COLUMNS["SKUS_ID"],
            value=str([x.sku.id for x in sku_grouped.skus]),
        )
        excel_client.draw_cell(
            row=beg_row, col=COLUMNS["BOILING_ID"], value=sku_grouped.boiling.id
        )
        if space_factor:
            cur_row += self.space_rows
        return cur_row

    def fill_mozzarella_sku_plan(self):
        self.skus_grouped = sorted(
            self.skus_grouped, key=lambda x: (x.boiling.is_lactose, x.boiling.percent)
        )
        sheet = self.wb[flask.current_app.config["SHEET_NAMES"]["schedule_plan"]]
        cur_row = 2
        is_lactose = False
        for sku_grouped in self.skus_grouped:
            if sku_grouped.boiling.is_lactose != is_lactose:
                cur_row += self.space_rows
            is_lactose = sku_grouped.boiling.is_lactose
            excel_client = ExcelBlock(sheet=sheet)
            cur_row = self.fill_skus(
                sku_grouped, excel_client, cur_row, is_lactose, True
            )

        for sheet_number, sheet_name in enumerate(self.wb.sheetnames):
            if sheet_number != 1:
                self.wb[sheet_name].views.sheetView[0].tabSelected = False
        self.wb.active = 1
        self.wb.save(self.filepath)

    def fill_ricotta_sku_plan(self):
        sheet = self.wb[flask.current_app.config["SHEET_NAMES"]["schedule_plan"]]
        cur_row = 2
        for sku_grouped in self.skus_grouped:
            excel_client = ExcelBlock(sheet=sheet)
            cur_row = self.fill_skus(sku_grouped, excel_client, cur_row, False)
            cur_row += self.space_rows

        for sheet_number, sheet_name in enumerate(self.wb.sheetnames):
            if sheet_number != 1:
                self.wb[sheet_name].views.sheetView[0].tabSelected = False
        self.wb.active = 1
        self.wb.save(self.filepath)

    def fill_butter_sku_plan(self):
        sheet = self.wb[flask.current_app.config["SHEET_NAMES"]["schedule_plan"]]
        cur_row = 2
        for sku_grouped in self.skus_grouped:
            excel_client = ExcelBlock(sheet=sheet)
            cur_row = self.fill_skus(sku_grouped, excel_client, cur_row, False, order=["Масло"])
            cur_row += self.space_rows

        for sheet_number, sheet_name in enumerate(self.wb.sheetnames):
            if sheet_number != 1:
                self.wb[sheet_name].views.sheetView[0].tabSelected = False
        self.wb.active = 1
        self.wb.save(self.filepath)

    def fill_milk_project_sku_plan(self):
        sheet = self.wb[flask.current_app.config["SHEET_NAMES"]["schedule_plan"]]
        cur_row = 2
        for sku_grouped in self.skus_grouped:
            excel_client = ExcelBlock(sheet=sheet)
            cur_row = self.fill_skus(sku_grouped, excel_client, cur_row, False,
                                     order=["Рикотта", "Качорикотта", "Четук", "Кавказский", "Черкесский"])
            cur_row += self.space_rows

        for sheet_number, sheet_name in enumerate(self.wb.sheetnames):
            if sheet_number != 1:
                self.wb[sheet_name].views.sheetView[0].tabSelected = False
        self.wb.active = 1
        self.wb.save(self.filepath)

    def fill_adygea_sku_plan(self):
        sheet = self.wb[flask.current_app.config["SHEET_NAMES"]["schedule_plan"]]
        cur_row = 2
        for sku_grouped in self.skus_grouped:
            excel_client = ExcelBlock(sheet=sheet)
            cur_row = self.fill_skus(sku_grouped, excel_client, cur_row, False,
                                     order=["Кавказский", "Черкесский", "Рикотта"])
            cur_row += self.space_rows

        for sheet_number, sheet_name in enumerate(self.wb.sheetnames):
            if sheet_number != 1:
                self.wb[sheet_name].views.sheetView[0].tabSelected = False
        self.wb.active = 1
        self.wb.save(self.filepath)

