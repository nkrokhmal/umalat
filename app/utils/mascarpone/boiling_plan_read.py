from collections import OrderedDict, defaultdict
from dataclasses import dataclass

import openpyxl
import pandas as pd
import numpy as np

from utils_ak.openpyxl import cast_workbook

from app.globals import db
from app.models.helpers import cast_model
from app.models.mascarpone import MascarponeSKU
from app.utils.mascarpone.utils import MascarponeBoilingsHandler
from app.utils.parse_remainings import cast_sku_name


class BoilingPlanReaderException(Exception):
    ...


@dataclass
class HugeBoiling:
    block_id: int = 0
    skus: list[dict] | None = None
    input_kg: int = 0
    output_kg: int = 0
    type: str | None = None
    washing: bool = False

    def __post_init__(self) -> None:
        self.skus = []


COLUMNS: dict[str, str] = {
    "Номер группы варок": "group_id",
    "Линия": "line",
    "Тип варки": "boiling_type",
    "SKU": "sku_name",
    "Кг": "kg",
    "Остатки": "leftovers",
    "Выход, кг": "output_kg",
    "Вход, кг": "input_kg",
    "Вставить мойку": "washing",
}


class BoilingPlanReader:
    def __init__(self, wb: openpyxl.Workbook, first_batches: dict | None = None) -> None:
        self.wb = cast_workbook(wb)
        self.first_batches: defaultdict = defaultdict(lambda: 1, first_batches)

    def _read_workbook(self) -> list[HugeBoiling]:
        ws = self.wb["План варок"]
        boilings: list[HugeBoiling] = []
        boiling = HugeBoiling()
        skus = db.session.query(MascarponeSKU).all()
        block_id = 1

        for i in range(2, 200):
            sku_name = ws.cell(i, 4).value
            match sku_name:
                case "-":
                    item = {column: ws.cell(i, j + 1).value for j, column in enumerate(COLUMNS.values())}
                    boiling.input_kg = item["input_kg"]
                    boiling.output_kg = item["output_kg"]
                    boiling.washing = item["washing"] == 1
                    boiling.block_id = block_id
                    boilings.append(boiling)

                    block_id += 1
                    boiling = HugeBoiling()
                case None | "":
                    continue
                case _:
                    item = {column: ws.cell(i, j + 1).value for j, column in enumerate(COLUMNS.values())}
                    sku = next((sku for sku in skus if sku.name == item["sku_name"]), None)
                    if sku is None:
                        raise BoilingPlanReaderException(
                            f"Неверно указано имя sku {item['sku_name']}, линия {i}, {item}"
                        )
                    boiling.type = self._get_group(sku)
                    boiling.skus.append(item)
        return boilings

    def _get_boilings(self, boilings: list[HugeBoiling]) -> pd.DataFrame:
        dfs = []
        group_dict = defaultdict(lambda: 1)
        for i, boiling in enumerate(boilings):
            df = pd.DataFrame(boiling.skus)
            df[["output_kg", "input_kg", "group_id", "block_id"]] = (
                boiling.output_kg,
                boiling.input_kg,
                group_dict[boiling.type],
                group_dict[boiling.type],
            )
            dfs.append(df)
            group_dict[boiling.type] += 1

        return pd.concat(dfs)

    @staticmethod
    def add_input_output_kg(df: pd.DataFrame) -> None:
        boilings = df["sku_name"].apply(lambda x: cast_sku_name(x).made_from_boilings[0])

        df["input_kg"] = boilings.apply(lambda x: x.input_kg)
        df["output_kg"] = boilings.apply(lambda x: x.input_kg * x.output_coeff)
        df["output_kg"] = np.round(df["output_kg"], -1)

    def _unwind_boilings(self, boilings: list[HugeBoiling]) -> pd.DataFrame:
        dfs = []
        group_id = 1
        for boiling in boilings:
            match boiling.type:
                case "cream":
                    df = pd.DataFrame(boiling.skus)
                    df[["output_kg", "input_kg", "group_id", "group", "block_id", "washing"]] = (
                        boiling.output_kg,
                        boiling.input_kg,
                        group_id,
                        boiling.type,
                        boiling.block_id,
                        boiling.washing,
                    )
                    dfs.append(df)
                    group_id += 1
                case _:
                    handler = MascarponeBoilingsHandler(check_boilings=True, use_boiling_weight=True)
                    handler.handle_group(boiling.skus, max_weight=0, weight_key="kg")
                    for i, group in enumerate(handler.boilings):
                        df = pd.DataFrame(group.skus)
                        if i == len(handler.boilings) - 1:
                            if df["kg"].sum() < 100:
                                df[["group_id", "block_id", "washing"]] = (
                                    group_id - 1,
                                    boiling.block_id,
                                    False,
                                )
                                self.add_input_output_kg(df)
                                dfs[-1] = pd.concat([dfs[-1], df])
                                continue

                        df[["group_id", "block_id", "washing"]] = (
                            group_id,
                            boiling.block_id,
                            False,
                        )
                        self.add_input_output_kg(df)
                        dfs.append(df)
                        group_id += 1

                    dfs[-1].iloc[-1, dfs[-1].columns.get_loc("washing")] = boiling.washing
        return pd.concat(dfs)

    def _saturate(self, df: pd.DataFrame) -> pd.DataFrame:
        df.index = range(len(df))
        df["sku"] = df["sku_name"].apply(lambda x: cast_model([MascarponeSKU], x))
        df["group"] = df["sku"].apply(lambda x: self._get_group(x))
        df["semifinished_group"] = df["group"].apply(lambda group: group if group != "cottage_cheese" else "robiola")
        df["batch_type"] = df["group"]
        df["batch_id"] = 0
        df["boiling"] = df["sku"].apply(lambda x: x.made_from_boilings[0])
        df["boiling_id"] = df["boiling"].apply(lambda boiling: boiling.id)
        return df

    @staticmethod
    def _get_group(sku: MascarponeSKU) -> str:
        match sku.group.name.lower():
            case "сливки":
                return "cream"
            case "кремчиз":
                return "cream_cheese"
            case "робиола":
                return "robiola"
            case "творожный":
                return "cottage_cheese"
            case "маскарпоне":
                return "mascarpone"
            case _:
                raise BoilingPlanReaderException(
                    """
                    Неизвестный тип максарпоне. В названии должен присутствовать один из типов:
                    Сливки, Кремчиз, Робиола, Творожный, Маскарпоне
                    """
                )

    def _set_batches(self, df: pd.DataFrame) -> pd.DataFrame:
        df["absolute_batch_id"] = df["batch_id"]

        # - Set batch_id

        group_counts = df.groupby(["block_id", "group"]).size().reset_index(name="count")
        idx = group_counts.groupby(["block_id"])["count"].transform(max) == group_counts["count"]
        most_popular_groups = group_counts[idx].drop("count", axis=1)
        most_popular_groups = most_popular_groups.drop_duplicates(subset=["block_id"])
        df = df.merge(most_popular_groups, on="block_id", suffixes=("", "_total_group"))
        df["batch_id"] = df.pop("block_id")

        # - Set absolute_batch_id

        batch_by_type = dict(self.first_batches)

        for batch_id, grp in df.groupby("batch_id"):
            batch_type = grp.iloc[0]["batch_type"]
            if batch_type == "cottage_cheese" or batch_type == "robiola":
                batch_type = "cream_cheese"
            df.loc[grp.index, "absolute_batch_id"] = batch_by_type[batch_type]
            batch_by_type[batch_type] += 1

        return df

    def parse(self, unwind: bool = False) -> pd.DataFrame:
        boilings = self._read_workbook()
        df = self._get_boilings(boilings) if unwind else self._unwind_boilings(boilings)
        df = self._saturate(df)
        df = self._set_batches(df)

        # update_absolute_batch_id(df, self.first_batches)
        df["boiling_id"] = df.pop("group_id")
        return df


__all__ = ["BoilingPlanReader", "BoilingPlanReaderException"]
