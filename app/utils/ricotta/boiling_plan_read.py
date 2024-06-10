from collections import defaultdict
from dataclasses import dataclass

import openpyxl
import pandas as pd

from utils_ak.openpyxl import cast_workbook

from app.globals import db
from app.lessmore.utils.get_repo_path import get_repo_path
from app.models.helpers import cast_model
from app.models.ricotta import RicottaSKU


class RicottaBoilingPlanReaderException(Exception):
    ...


@dataclass
class HugeBoiling:
    skus: list[dict] | None = None
    output_kg: int = 0
    boiling_num: float = 1.0

    def __post_init__(self) -> None:
        self.skus = []


COLUMNS: dict[str, str] = {
    "Номер группы варок": "group_id",
    "Тип варки": "boiling_type",
    "Выход с одной варки, кг": "output_kg_per_one",
    "SKU": "sku_name",
    "КГ": "kg",
    "Остатки": "leftovers",
    "Суммарный вес  сыворотки": "sum_weight_kg",
    "Вес на выходе одной варки": "output_kg",
    "Количество флокуляторов": "floculators_num",
    "Количество варок": "boilings_num",
}


class BoilingPlanReader:
    def __init__(self, wb: openpyxl.Workbook, first_batches: dict | None = None) -> None:
        self.wb = cast_workbook(wb)
        self.first_batches: defaultdict = defaultdict(lambda: 1, first_batches)

    @staticmethod
    def _validate(item: dict) -> None:
        if item["boilings_num"] == 1:
            if abs(item["leftovers"]) > 0.1 * item["output_kg"]:
                raise RicottaBoilingPlanReaderException("Остатки превышают 10 процентов веса варки")
        else:
            if abs(item["leftovers"]) > 1:
                raise RicottaBoilingPlanReaderException("Остатки в мульти варках должны всегда быть 0")

    def _read_workbook(self) -> list[HugeBoiling]:
        ws = self.wb["План варок"]
        boilings: list[HugeBoiling] = []
        boiling = HugeBoiling()
        skus = db.session.query(RicottaSKU).all()

        for i in range(2, 200):
            sku_name = ws.cell(i, 4).value
            match sku_name:
                case "-":
                    item = {column: ws.cell(i, j + 1).value for j, column in enumerate(COLUMNS.values())}
                    self._validate(item)
                    boiling.boiling_num = item["boilings_num"]
                    boiling.output_kg = item["output_kg"]
                    boilings.append(boiling)

                    boiling = HugeBoiling()
                case None | "":
                    continue
                case _:
                    item = {column: ws.cell(i, j + 1).value for j, column in enumerate(COLUMNS.values())}
                    sku = next((sku for sku in skus if sku.name == item["sku_name"]), None)
                    if sku is None:
                        raise RicottaBoilingPlanReaderException(f"Неверно указано имя sku {item['sku_name']}")
                    boiling.skus.append(item)
        return boilings

    def _unwind_boilings(self, boilings: list[HugeBoiling]) -> pd.DataFrame:
        dfs = []
        group_id = 0
        for boiling in boilings:
            full_boilings = int(boiling.boiling_num)
            df = pd.DataFrame(boiling.skus)

            if len(df) > 1 and full_boilings > 1:
                skus = " ,\n".join(df.sku_name.to_list())
                raise RicottaBoilingPlanReaderException(
                    f"""
                    В плане варок есть варка с SKU
                    
                    {skus}
                    
                    и количеством варок {full_boilings}. 
                    
                    Разрешается иметь только одну SKU, когда количество варок больше 1. 
                    Пожалуйста, исправьте файл варок! 
                """
                )

            for i in range(full_boilings):
                df = pd.DataFrame(boiling.skus)
                df[["output_kg", "group_id", "floculators_num", "boilings_num", "sum_weight_kg"]] = (
                    boiling.output_kg,
                    group_id,
                    2,
                    1,
                    6500 * 2,
                )
                if full_boilings > 1:
                    df["kg"] = df["output_kg"]
                group_id += 1
                dfs.append(df)

            if boiling.boiling_num - full_boilings > 0.1:
                df = pd.DataFrame(boiling.skus)
                df[["output_kg", "group_id", "floculators_num", "boilings_num", "sum_weight_kg"]] = (
                    boiling.output_kg,
                    group_id,
                    1,
                    0.5,
                    6500,
                )
                group_id += 1
                dfs.append(df)
        return pd.concat(dfs)

    def _saturate(self, df: pd.DataFrame) -> pd.DataFrame:
        df.index = range(len(df))
        df["sku"] = df["sku_name"].apply(lambda x: cast_model([RicottaSKU], x))
        df["boiling"] = df["sku"].apply(lambda x: x.made_from_boilings[0])
        df["boiling_id"] = df["boiling"].apply(lambda boiling: boiling.id)
        df["batch_type"] = "ricotta"
        df.drop(["leftovers", "output_kg"], axis=1)

        return df

    def _set_batches(self, df: pd.DataFrame) -> pd.DataFrame:
        for _, grp in df.groupby("group_id"):
            group = grp.iloc[0][["batch_type"]]
            df.loc[grp.index, "batch_id"] = self.first_batches[group.iloc[0]]
            self.first_batches[group[0]] += 1

        df["absolute_batch_id"] = df["batch_id"]
        return df

    def parse(self) -> pd.DataFrame:
        boilings = self._read_workbook()
        df = self._unwind_boilings(boilings)
        df = self._saturate(df)
        df = self._set_batches(df)
        return df


def test():

    # - Configure pandas

    pd.set_option("display.max_rows", 500)
    pd.set_option("display.max_columns", 500)
    pd.set_option("display.width", 1000)

    # - Parse and read dataframe

    print(
        BoilingPlanReader(
            openpyxl.load_workbook(str(get_repo_path() / "app/data/tests/ricotta/boiling.xlsx"), data_only=True),
            first_batches={"ricotta": 1},
        ).parse()
    )


__all__ = ["BoilingPlanReader", "RicottaBoilingPlanReaderException"]

if __name__ == "__main__":
    test()
