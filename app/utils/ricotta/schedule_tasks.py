import os
from flask import current_app
from openpyxl.styles import Alignment
from collections import namedtuple
import math
from app.enum import LineName
from app.utils.features.openpyxl_wrapper import ExcelBlock
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


Cell = namedtuple("Cell", "col, col_name")
COLOR = "#dce6f2"
COLUMNS = {
    "index": Cell(column_index_from_string("B"), "B"),
    "sku": Cell(column_index_from_string("C"), "C"),
    "boxes": Cell(column_index_from_string("I"), "I"),
    "kg": Cell(column_index_from_string("J"), "J"),
    "boxes_count": Cell(column_index_from_string("K"), "K"),
    "priority": Cell(column_index_from_string("L"), "L"),
}

FONTS = {"header": 12, "title": 10, "body": 9}

DIMENSIONS = {"header": 30, "title": 28, "body": 22}


def draw_header(excel_client, date, cur_row, task_name, is_boiling=None):
    excel_client.colour = None
    excel_client.sheet.column_dimensions[COLUMNS["sku"].col_name].width = 5 * 5
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    excel_client.font_size = FONTS["header"]
    excel_client.raw_dimension(cur_row, DIMENSIONS["header"])
    excel_client.merge_cells(
        beg_col=COLUMNS["index"].col,
        beg_row=cur_row,
        end_col=COLUMNS["priority"].col,
        end_row=cur_row,
        value=task_name,
        alignment=alignment,
        is_bold=True,
        font_size=FONTS["header"],
    )
    cur_row += 1

    excel_client.raw_dimension(cur_row, DIMENSIONS["header"])
    excel_client.merge_cells(
        beg_col=COLUMNS["index"].col,
        beg_row=cur_row,
        end_col=COLUMNS["priority"].col,
        end_row=cur_row,
        value=date.date(),
        alignment=alignment,
        is_bold=True,
        font_size=FONTS["header"],
    )
    cur_row += 1

    excel_client.font_size = FONTS["title"]
    excel_client.raw_dimension(cur_row, DIMENSIONS["title"])
    excel_client.colour = COLOR[1:]
    excel_client.draw_cell(
        col=COLUMNS["index"].col,
        row=cur_row,
        value="Номер {}".format(is_boiling) if is_boiling is not None else "Номер",
        alignment=alignment,
    )
    excel_client.merge_cells(
        beg_col=COLUMNS["sku"].col,
        beg_row=cur_row,
        end_col=COLUMNS["boxes"].col - 1,
        end_row=cur_row,
        value="Номенклатура",
        alignment=alignment,
        font_size=FONTS["header"],
    )
    excel_client.draw_cell(
        col=COLUMNS["boxes"].col,
        row=cur_row,
        value="Вложение коробок",
        alignment=alignment,
    )
    excel_client.draw_cell(
        col=COLUMNS["kg"].col, row=cur_row, value="Вес, кг", alignment=alignment
    )
    excel_client.draw_cell(
        col=COLUMNS["boxes_count"].col,
        row=cur_row,
        value="Кол-во коробок, шт",
        alignment=alignment,
    )
    excel_client.draw_cell(
        col=COLUMNS["priority"].col,
        row=cur_row,
        value="В первую очередь",
        alignment=alignment,
    )
    cur_row += 1
    return cur_row, excel_client


def draw_task_new(excel_client, df, date, cur_row, task_name, batch_number):
    cur_row, excel_client = draw_header(excel_client, date, cur_row, task_name, "варки")
    for boiling_group_id, grp in df.groupby("boiling_id"):
        for i, row in grp.iterrows():
            excel_client.raw_dimension(cur_row, DIMENSIONS["body"])
            excel_client.font_size = FONTS["body"]
            excel_client.colour = COLOR[1:]

            excel_client.draw_cell(
                col=COLUMNS["index"].col,
                row=cur_row,
                value=boiling_group_id + batch_number - 1,
            )
            excel_client.colour = None
            excel_client.merge_cells(
                beg_col=COLUMNS["sku"].col,
                beg_row=cur_row,
                end_col=COLUMNS["boxes"].col - 1,
                end_row=cur_row,
                value=row["sku_name"],
            )
            excel_client.draw_cell(
                col=COLUMNS["boxes"].col, row=cur_row, value=row["sku"].boxes
            )

            if row["sku"].group.name != "Качокавалло":
                kg = round(row["kg"])
                boxes_count = math.ceil(
                    1000 * row["kg"] / row["sku"].boxes / row["sku"].weight_netto
                )
            else:
                kg = ""
                boxes_count = ""

            excel_client.draw_cell(col=COLUMNS["kg"].col, row=cur_row, value=kg)
            excel_client.draw_cell(
                col=COLUMNS["boxes_count"].col, row=cur_row, value=boxes_count
            )
            excel_client.draw_cell(col=COLUMNS["priority"].col, row=cur_row, value="")
            cur_row += 1

        excel_client.colour = COLOR[1:]
        excel_client.draw_cell(col=COLUMNS["index"].col, row=cur_row, value="")
        excel_client.merge_cells(
            beg_col=COLUMNS["sku"].col,
            beg_row=cur_row,
            end_col=COLUMNS["priority"].col,
            end_row=cur_row,
            value="",
        )
        cur_row += 1
    return cur_row


def schedule_task_boilings(wb, df, date, batch_number):
    df_copy = df.copy()
    sheet_name = "Печать заданий"
    ricotta_task_name = "Задание на упаковку Рикоттного цеха"

    cur_row = 2
    space_row = 4

    wb.create_sheet(sheet_name)
    excel_client = ExcelBlock(wb[sheet_name], font_size=9)

    cur_row = draw_task_new(
        excel_client,
        df_copy,
        date,
        cur_row,
        ricotta_task_name,
        batch_number,
    )
    cur_row += space_row
    return wb
