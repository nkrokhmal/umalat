from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import openpyxl
from copy import copy


class ExcelBlock:
    def __init__(self, sheet, colour='FFFFFF', font_size=8):
        self.sheet = sheet
        self.colour = colour
        self.font_size = font_size

    def raw_dimension(self, row, row_height):
        self.sheet.row_dimensions[row].height = row_height

    def default_colour(self, row, col, set_colour):
        if set_colour:
            if (self.colour != 'FFFFFF') and (self.colour != None):
                self.sheet.cell(row, col).fill = PatternFill("solid", fgColor=self.colour)

    def default_font(self, row, col, set_font):
        if set_font:
            self.sheet.cell(row, col).font = Font(size=self.font_size)

    def default_border(self, row, col):
        self.sheet.cell(row, col).border = Border(left=Side(style='thin'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin'))

    def draw_cell(self, row, col, value, alignment=None, set_colour=True, set_font=True, set_border=True,
                  font_size=None, is_bold=False):
        self.sheet.cell(row, col).value = value
        self.default_colour(row, col, set_colour)
        self.default_font(row, col, set_font)

        if set_border:
            self.default_border(row, col)

        if font_size:
            self.sheet.cell(row, col).font = Font(size=self.font_size, bold=is_bold)

        if alignment is not None:
            self.sheet.cell(row, col).alignment = alignment

    def draw_row(self, row, values, cols=None, alignment=None, set_colour=True, set_font=True, set_border=True):
        for i, value in enumerate(values):
            col = cols[i] if cols else i + 1
            self.draw_cell(row, col, value, alignment=alignment, set_colour=set_colour,
                           set_font=set_font, set_border=set_border)

    def merge_cells(self, beg_row, end_row, beg_col, end_col, value, alignment=None, set_colour=True, set_font=True,
                    set_border=True, font_size=None, is_bold=False):
        self.sheet.merge_cells(start_row=beg_row,
                               end_row=end_row,
                               start_column=beg_col,
                               end_column=end_col)
        self\
            .default_colour(beg_row, beg_col, set_colour)
        self.default_font(beg_row, beg_col, set_font)
        self.draw_cell(row=beg_row, col=beg_col, value=value, alignment=alignment, set_colour=set_colour,
                       set_font=set_font, set_border=set_border, font_size=font_size, is_bold=is_bold)
        if set_border:
            for row in range(beg_row, end_row + 1):
                for col in range(beg_col, end_col + 1):
                    self.default_border(row, col)


def merge_workbooks(wb1, wb2):
    for i, sheet_name in enumerate(wb1.sheetnames):
        ws1 = wb1[sheet_name]
        max_row = ws1.max_row
        max_col = ws1.max_column
        ws2 = wb2.create_sheet(sheet_name, i)

        for i in range(1, max_col + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].hidden = ws1.column_dimensions[
                openpyxl.utils.get_column_letter(i)].hidden
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ws1.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width

        for i in range(1, max_row + 1):
            for j in range(1, max_col + 1):
                ws2.cell(row=i, column=j).value = ws1.cell(row=i, column=j).value
                if ws1.cell(row=i, column=j).has_style:
                    ws2.cell(row=i, column=j).fill = copy(ws1.cell(row=i, column=j).fill)
                    ws2.cell(row=i, column=j).font = copy(ws1.cell(row=i, column=j).font)
                    ws2.cell(row=i, column=j).border = copy(ws1.cell(row=i, column=j).border)
                    ws2.cell(row=i, column=j).alignment = copy(ws1.cell(row=i, column=j).alignment)
    return ws2
