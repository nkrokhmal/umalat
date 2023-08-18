from collections import namedtuple

from openpyxl.styles import Alignment
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from app.enum import LineName
from app.imports.runtime import *
from app.utils.features.openpyxl_wrapper import ExcelBlock


Cell = namedtuple("Cell", "col, col_name")
COLOR = "#dce6f2"
COLOR_PACKING = "#E0E0E0"
COLUMNS = {
    "index": Cell(column_index_from_string("B"), "B"),
    "sku": Cell(column_index_from_string("C"), "C"),
    "boxes": Cell(column_index_from_string("I"), "I"),
    "kg": Cell(column_index_from_string("J"), "J"),
    "boxes_count": Cell(column_index_from_string("K"), "K"),
    "priority": Cell(column_index_from_string("L"), "L"),
    "code": Cell(column_index_from_string("M"), "M"),
    "end": Cell(column_index_from_string("N"), "N"),
}

FONTS = {"header": 12, "title": 10, "body": 9}

DIMENSIONS = {"header": 30, "title": 28, "body": 22}


def draw_header(excel_client, date, cur_row, task_name, is_boiling=None):
    excel_client.colour = None
    excel_client.sheet.column_dimensions[COLUMNS["sku"].col_name].width = 5 * 5
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    excel_client.font_size = FONTS["header"]
    excel_client.raw_dimension(cur_row, DIMENSIONS["header"])
    excel_client.merge_cells(
        beg_col=COLUMNS["index"].col,
        beg_row=cur_row,
        end_col=COLUMNS["end"].col,
        end_row=cur_row,
        value=task_name,
        alignment=alignment,
        is_bold=True,
        font_size=FONTS["header"],
    )
    cur_row += 1

    excel_client.raw_dimension(cur_row, DIMENSIONS["header"])
    excel_client.merge_cells(
        beg_col=COLUMNS["index"].col,
        beg_row=cur_row,
        end_col=COLUMNS["end"].col,
        end_row=cur_row,
        value=date.date(),
        alignment=alignment,
        is_bold=True,
        font_size=FONTS["header"],
    )
    cur_row += 1

    excel_client.font_size = FONTS["title"]
    excel_client.raw_dimension(cur_row, DIMENSIONS["title"])
    excel_client.colour = COLOR[1:]
    excel_client.draw_cell(
        col=COLUMNS["index"].col,
        row=cur_row,
        value="Номер {}".format(is_boiling) if is_boiling is not None else "Номер",
        alignment=alignment,
    )
    excel_client.merge_cells(
        beg_col=COLUMNS["sku"].col,
        beg_row=cur_row,
        end_col=COLUMNS["boxes"].col - 1,
        end_row=cur_row,
        value="Номенклатура",
        alignment=alignment,
        font_size=FONTS["header"],
    )
    excel_client.draw_cell(
        col=COLUMNS["boxes"].col,
        row=cur_row,
        value="Вложение коробок",
        alignment=alignment,
    )
    excel_client.draw_cell(col=COLUMNS["kg"].col, row=cur_row, value="Вес, кг", alignment=alignment)
    excel_client.draw_cell(
        col=COLUMNS["boxes_count"].col,
        row=cur_row,
        value="Кол-во коробок, шт",
        alignment=alignment,
    )
    excel_client.draw_cell(
        col=COLUMNS["priority"].col,
        row=cur_row,
        value="В первую очередь",
        alignment=alignment,
    )

    excel_client.merge_cells(
        beg_col=COLUMNS["code"].col,
        beg_row=cur_row,
        end_col=COLUMNS["end"].col,
        end_row=cur_row,
        value="Код",
        alignment=alignment,
        font_size=FONTS["header"],
    )
    cur_row += 1
    return cur_row, excel_client


def draw_schedule_raw(excel_client, cur_row, values, color=None):
    excel_client.raw_dimension(cur_row, DIMENSIONS["body"])
    excel_client.font_size = FONTS["body"]
    excel_client.colour = COLOR[1:]

    excel_client.draw_cell(
        col=COLUMNS["index"].col,
        row=cur_row,
        value=values[0],
    )
    excel_client.colour = color[1:] if color else None
    excel_client.merge_cells(
        beg_col=COLUMNS["sku"].col,
        beg_row=cur_row,
        end_col=COLUMNS["boxes"].col - 1,
        end_row=cur_row,
        value=values[1],
    )
    excel_client.draw_cell(col=COLUMNS["boxes"].col, row=cur_row, value=values[2])
    excel_client.draw_cell(col=COLUMNS["kg"].col, row=cur_row, value=values[3])
    excel_client.draw_cell(col=COLUMNS["boxes_count"].col, row=cur_row, value=values[4])
    excel_client.draw_cell(col=COLUMNS["priority"].col, row=cur_row, value="")
    excel_client.merge_cells(
        beg_col=COLUMNS["code"].col,
        beg_row=cur_row,
        end_col=COLUMNS["end"].col,
        end_row=cur_row,
        value=values[5],
    )

    excel_client.colour = None
    cur_row += 1
    return excel_client, cur_row


def draw_blue_line(excel_client, cur_row):
    excel_client.colour = COLOR[1:]
    excel_client.draw_cell(col=COLUMNS["index"].col, row=cur_row, value="")
    excel_client.merge_cells(
        beg_col=COLUMNS["sku"].col,
        beg_row=cur_row,
        end_col=COLUMNS["end"].col,
        end_row=cur_row,
        value="",
    )
    return excel_client
