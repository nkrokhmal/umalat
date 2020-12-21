from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


class ExcelBlock:
    def __init__(self, sheet, colour='#FFFFFF', font_size=9, row_height=14):
        self.sheet = sheet
        self.colour = colour
        self.font_size = font_size
        self.row_height = row_height
        for i in range(2, 100):
            self.sheet.row_dimensions[i].height = self.row_height

    def default_colour(self, row, col, set_colour):
        if set_colour:
            self.sheet.cell(row, col).fill = PatternFill("solid", fgColor=self.colour)

    def default_font(self, row, col, set_font):
        if set_font:
            self.sheet.cell(row, col).font = Font(size=self.font_size)

    def default_border(self, row, col):
        self.sheet.cell(row, col).border = Border(left=Side(style='thin'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin'))

    def cell_value(self, row, col, value, alignment=None, set_colour=True, set_font=True, set_border=True):
        self.sheet.cell(row, col).value = value
        self.default_colour(row, col, set_colour)
        self.default_font(row, col, set_font)

        if set_border:
            self.default_border(row, col)
        if alignment is not None:
            self.sheet.cell(row, col).alignment = alignment
        else:
            self.sheet.cell(row, col).alignment = Alignment(wrapText=True)

    def merge_cells(self, beg_row, end_row, beg_col, end_col, value, alignment=None, set_colour=True, set_font=True, set_border=True):
        self.sheet.merge_cells(start_row=beg_row,
                               end_row=end_row,
                               start_column=beg_col,
                               end_column=end_col)
        self.cell_value(row=beg_row, col=beg_col, value=value, alignment=alignment, set_colour=set_colour, set_font=set_font, set_border=set_border)
        self.default_colour(beg_row, beg_col, set_colour)
        self.default_font(beg_row, beg_col, set_font)
        if set_border:
            for row in range(beg_row + 1, end_row + 1):
                self.sheet.cell(end_row, end_col).border = Border(left=Side(style='thin'),
                                                                  right=Side(style='thin'))
            if end_row != beg_row:
                self.sheet.cell(end_row, end_col).border = Border(bottom=Side(style='thin'))







