import numpy as np
import openpyxl as opx
import pandas as pd

from utils_ak.code_block import code
from utils_ak.code_block.code import code
from utils_ak.numeric.types import is_int_like
from utils_ak.openpyxl.openpyxl_tools import cast_workbook

from app.enum import LineName
from app.models import Line, MozzarellaSKU, cast_model, cast_mozzarella_boiling, cast_mozzarella_form_factor
from app.scheduler.boiling_plan import update_absolute_batch_id
from app.scheduler.mozzarella.boiling_plan.saturate import saturate_boiling_plan
from config import config


def read_sheet(wb, sheet_name, default_boiling_volume=1000, sheet_number=1):
    ws = wb[sheet_name]
    values = []

    # collect header
    header = [ws.cell(1, i).value for i in range(1, 100) if ws.cell(1, i).value]

    for i in range(2, 200):
        if not ws.cell(i, 2).value:
            continue

        values.append([ws.cell(i, j).value for j in range(1, len(header) + 1)])

    df = pd.DataFrame(values, columns=header)
    df = df[
        [
            "Тип варки",
            "Объем варки",
            "SKU",
            "КГ",
            "Форм фактор плавления",
            "Номер команды",
            "Конфигурация варки",
            "Вес варки",
            "Мойка",
        ]
    ]
    df.columns = [
        "boiling_params",
        "boiling_volume",
        "sku",
        "kg",
        "bff",
        "packing_team_id",
        "configuration",
        "total_volume",
        "cleaning",
    ]

    # fill group id
    df["group_id"] = (df["boiling_params"] == "-").astype(int).cumsum() + 1

    with code("Convert total_volume to float safely"):

        def _cast_float(obj):
            try:
                return float(obj)
            except:
                return np.nan

        df["total_volume"] = df["total_volume"].apply(_cast_float)
    # fill total_volume
    df["total_volume"] = np.where(
        (df["sku"] == "-") & (df["total_volume"].isnull()),
        default_boiling_volume,
        df["total_volume"],
    )
    df["total_volume"] = df["total_volume"].fillna(method="bfill")

    # fill cleaning
    df["cleaning"] = np.where((df["sku"] == "-") & (df["cleaning"].isnull()), "", df["cleaning"])
    df["cleaning"] = df["cleaning"].fillna(method="bfill")
    df["cleaning"] = df["cleaning"].apply(
        lambda cleaning_type: {
            "Короткая мойка": "short",
            "Длинная мойка": "full",
        }.get(cleaning_type, "")
    )

    # fill configuration
    def format_configuration(value):
        if is_int_like(value):
            return str(int(value))
        elif value is None:
            return None
        elif np.isnan(value):
            return None
        elif isinstance(value, str):
            assert "," not in value, "Группы варок не поддерживаются в моцаррельном цеху."
            return value
        else:
            raise AssertionError("Unknown format")

    df["configuration"] = df["configuration"].apply(format_configuration)
    df["configuration"] = np.where(
        (df["sku"] == "-") & (df["configuration"].isnull()),
        "8000",
        df["configuration"],
    )
    df["configuration"] = df["configuration"].fillna(method="bfill")

    # remove separators and empty lines
    df = df[df["sku"] != "-"]
    df = df[~df["kg"].isnull()]
    df["sku_obj"] = df["sku"].apply(lambda sku: cast_model(MozzarellaSKU, sku))
    df["sku_obj"] = df["sku_obj"].apply(lambda x: x.line.name)

    # add line name to boiling_params
    df["boiling_params"] = df.apply(lambda row: row["sku_obj"] + "," + row["boiling_params"], axis=1)
    df["sheet"] = sheet_number

    return df


def update_boiling_plan(dfs, normalization, saturate, validate=True):
    df = pd.concat(dfs).reset_index(drop=True)
    if len(df) == 0:
        return pd.DataFrame()

    with code("Fix group ids: first water -> then salt"):
        cur_group_id = 1
        for line_name in [LineName.WATER, LineName.SALT]:
            for ind, grp in df[df["sku_obj"] == line_name].groupby("group_id"):
                df.loc[grp.index, "group_id"] = cur_group_id
                cur_group_id += 1

    df["sku"] = df["sku"].apply(lambda sku: cast_model(MozzarellaSKU, sku))

    df["boiling"] = df["boiling_params"].apply(cast_mozzarella_boiling)

    # set boiling form factors
    df["ff"] = df["sku"].apply(lambda sku: sku.form_factor)

    # remove Терка from form_factors
    df["_bff"] = df["ff"].apply(lambda ff: ff if "Терка" not in ff.name else None)

    # fill Терка empty form factor values
    for idx, grp in df.copy().groupby("group_id"):
        if grp["_bff"].isnull().all():
            # take from bff input if not specified
            df.loc[grp.index, "_bff"] = cast_mozzarella_form_factor(config.DEFAULT_RUBBER_FORM_FACTOR)
        else:
            filled_grp = grp.copy()
            filled_grp = filled_grp.fillna(method="ffill")
            filled_grp = filled_grp.fillna(method="bfill")
            df.loc[grp.index, "_bff"] = filled_grp["_bff"]
    df["bff"] = df["_bff"]

    # validate single boiling
    for idx, grp in df.groupby("group_id"):
        assert len(grp["boiling"].unique()) == 1, "В одной группе варок должен быть только один тип варки."

    df["original_kg"] = df["kg"]

    assert (
        not df["total_volume"].isnull().any()
    ), "Ошибка в чтении плана варок. Если вы работаете с файлом расписания, убедитесь что вы его сохранили перед тем, как залить на сайт"

    # validate kilograms
    if validate:
        for idx, grp in df.groupby("group_id"):
            if abs(grp["kg"].sum() - grp.iloc[0]["total_volume"]) / grp.iloc[0]["total_volume"] > 0.05:
                raise AssertionError("Одна из групп варок имеет неверное количество килограмм.")
            else:
                if normalization:
                    if abs(grp["kg"].sum() - grp.iloc[0]["total_volume"]) > 1e-5:
                        df.loc[grp.index, "kg"] *= (
                            grp.iloc[0]["total_volume"] / grp["kg"].sum()
                        )  # scale to total_volume
                    else:
                        # all fine
                        pass

    df = df[
        [
            "group_id",
            "sku",
            "kg",
            "original_kg",
            "packing_team_id",
            "boiling",
            "bff",
            "configuration",
            "cleaning",
            "sheet",
            "total_volume",
        ]
    ]

    df = df.reset_index(drop=True)

    if saturate:
        df = saturate_boiling_plan(df)
    return df


def read_boiling_plan(wb_obj, saturate=True, normalization=True, validate=True, first_batch_ids=None):
    """
    :param wb_obj: str or openpyxl.Workbook
    :return: pd.DataFrame(columns=['id', 'boiling', 'sku', 'kg'])
    """
    first_batch_ids = first_batch_ids or {"mozzarella": 1}
    wb = cast_workbook(wb_obj)

    dfs = []

    sheet_names = ["Вода", "Соль", "План варок"]
    sheet_names = [sheet_name for sheet_name in sheet_names if sheet_name in wb.sheetnames]

    for i, ws_name in enumerate(sheet_names):
        if ws_name in ["Вода", "Соль"]:
            line = cast_model(Line, LineName.WATER) if ws_name == "Вода" else cast_model(Line, LineName.SALT)
            default_boiling_volume = line.output_kg
        else:
            default_boiling_volume = None

        df = read_sheet(wb, ws_name, default_boiling_volume=default_boiling_volume, sheet_number=i)
        dfs.append(df)

    df = update_boiling_plan(dfs, normalization, saturate, validate)
    df["packing_team_id"] = df["packing_team_id"].apply(lambda x: int(x))

    # batch_id and boiling_id are the same as group_id
    df["batch_id"] = df["group_id"]
    df["boiling_id"] = df["group_id"]
    df["batch_type"] = "mozzarella"
    df = update_absolute_batch_id(df, first_batch_ids)
    return df


def read_additional_packing(wb_obj):
    wb = cast_workbook(wb_obj)
    ws = wb["Дополнительная фасовка"]

    values = []
    for i in range(2, 10):
        if not ws.cell(i, 2).value:
            continue

        values.append([ws.cell(i, j).value for j in range(1, 3)])

    df = pd.DataFrame(values, columns=["sku", "kg"])
    df["sku_obj"] = df["sku"].apply(lambda sku: cast_model(MozzarellaSKU, sku))
    df["kg"] = -df["kg"]
    df = df[df["kg"] > 0]
    return df


def cast_boiling_plan(boiling_plan_obj, *args, **kwargs):
    if isinstance(boiling_plan_obj, (str, opx.Workbook)):
        return read_boiling_plan(boiling_plan_obj, *args, **kwargs)
    elif isinstance(boiling_plan_obj, pd.DataFrame):
        return boiling_plan_obj
    else:
        raise Exception("Unknown boiling plan type")
