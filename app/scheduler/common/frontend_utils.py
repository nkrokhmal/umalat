from loguru import logger
from openpyxl.styles import Side
from openpyxl.styles.borders import BORDER_THIN
from utils_ak.block_tree.validation.is_disjoint import is_disjoint
from utils_ak.color.color import cast_color
from utils_ak.openpyxl.openpyxl_tools import (
    cast_worksheet,
    draw_merged_cell,
    init_workbook,
    set_active_sheet,
    set_border_grid,
    set_dimensions,
    set_zoom,
)
from utils_ak.os.os_tools import makedirs, open_file_in_os

from app.scheduler.common.time_utils import cast_t


def fill_grid(ws):
    set_border_grid(ws, 1, 1, ws.max_column, ws.max_row, Side(border_style=BORDER_THIN))


def prepare_schedule_worksheet(ws_obj):
    ws = cast_worksheet(ws_obj)
    set_zoom(ws, 55)
    set_dimensions(ws, "column", range(1, 5), 21)
    set_dimensions(ws, "column", range(5, 288 * 2), 2.4)
    set_dimensions(ws, "row", range(1, 220), 25)
    return ws


def init_schedule_workbook(wb=None):
    if not wb:
        wb = init_workbook(["Расписание"])

    if "Расписание" not in wb.sheetnames:
        wb.create_sheet("Расписание")

    prepare_schedule_worksheet(cast_worksheet((wb, "Расписание")))
    return wb


def draw_schedule(schedule, style, O=None, fn=None, wb=None, debug=False, init=True):
    if init:
        wb = init_schedule_workbook(wb)
    logger.info("Finished init")
    ws = cast_worksheet((wb, "Расписание"))
    set_active_sheet(wb, "Расписание")
    O = O or [0, 0]  # initial coordinates

    # update styles
    for b in schedule.iter():
        block_style = style.get(b.props["cls"])

        if block_style:
            block_style = {k: v(b) if callable(v) else v for k, v in block_style.items()}
            b.props.update(**block_style)

    schedule.props.update(index_width=4)
    #
    # # - Check for conflicts (not tested extensively, hence for debugging only for now)
    # # [AttributeError: 'MergedCell' object attribute 'value' is read-only]
    #
    # leaves = [
    #     b for b in schedule.iter() if b.is_leaf() and b.props.get("visible", True)
    # ]
    # for i, b2 in enumerate(leaves):
    #     if b2.size[0] < 0 or b2.size[1] < 0:
    #         print("Block size is negative")
    #         print(b2)
    #         raise Exception(
    #             "Ошибка в построении расписания, произошли накладки одних блоков на другие. Такого быть не должно, нужно обращатсья к разработчикам."
    #         )
    #     for b1 in leaves[:i]:
    #
    #         if not is_disjoint(b1, b2):
    #             print("Block conflict")
    #             print(b1, b1.props["boiling_id"])
    #             print(b2, b2.props["boiling_id"])
    #             raise Exception(
    #                 "Ошибка в построении расписания, произошли накладки одних блоков на другие. Такого быть не должно, нужно обращатсья к разработчикам."
    #             )

    # - Draw schedule

    for b in schedule.iter():
        if b.is_leaf() and b.props.get("visible", True):
            if b.size[0] == 0 or b.size[1] == 0:
                continue

            text = b.props.get("text", "")
            color = cast_color(b.props.get("color", "white"))

            try:
                text = text.format(**b.props.all())
                text = text.replace("<", "{")
                text = text.replace(">", "}")
                text = eval(f"f{text!r}")

                # print(b.props['cls'], b.x, b.y)

                x1 = b.x[0]
                if "start_time" in b.props.all():
                    x1 -= cast_t(b.props["start_time"])  # shift of timeline
                x1 += b.props["index_width"]  # first index columns
                x1 += 1  # indexing starts with 1 in excel

                bold = b.props["bold"]
                font_size = b.props.get("font_size", 12)
                # print(b.props['cls'], x1, b.x[1], b.size[0], b.size[1])

                try:
                    draw_merged_cell(
                        ws,
                        x1 + O[0],
                        b.x[1] + O[1],
                        b.size[0],
                        b.size[1],
                        text,
                        color,
                        bold=bold,
                        border=b.props.relative_props.get("border", {"border_style": "thin", "color": "000000"}),
                        text_rotation=b.props.get("text_rotation"),
                        font_size=font_size,
                        alignment="center",
                    )
                except AttributeError as e:
                    if "'MergedCell' object attribute 'value' is read-only" in str(e):
                        logger.exception("Block conflict during drawing")
                    if not debug:
                        raise
            except:
                logger.error("Failed to draw block", b=b, x=(x1, b.x[1]), size=b.size)
                logger.error("Relative props", props=b.props.relative_props)
                raise Exception(
                    "Ошибка в построении расписания. Такого быть не должно, нужно обращаться к разработчикам."
                )

    if fn:
        wb.save(fn)

    return wb


def draw_excel_frontend(frontend, style, O=None, open_file=False, fn=None, wb=None, init=True):
    wb = draw_schedule(frontend, style, O=O, wb=wb, init=init)

    if fn:
        makedirs(fn)
        wb.save(fn)

        if open_file:
            open_file_in_os(fn)

    return wb
