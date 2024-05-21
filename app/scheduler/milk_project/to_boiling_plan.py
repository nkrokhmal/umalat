from typing import Union

import pandas as pd

from openpyxl import Workbook
from utils_ak.builtin import delistify
from utils_ak.code_block import code
from utils_ak.code_block.code import code
from utils_ak.openpyxl.openpyxl_tools import cast_workbook

from app.lessmore.utils.get_repo_path import get_repo_path
from app.models import MilkProjectSKU, cast_model
from app.scheduler.boiling_plan_like import BoilingPlanLike
from app.scheduler.calc_absolute_batch_id import calc_absolute_batch_id


def to_boiling_plan(boiling_plan: BoilingPlanLike, first_batch_ids_by_type={"milk_project": 1}):
    """Считать файл плана варок в датафрейм

    Может читать и файл расписания, т.к. там там обычно есть лист с планом варок

    Parameters
    ----------
    boiling_plan_source : str or openpyxl.Workbook or pd.DataFrame
        Путь к файлу плана варок или сам файл

    Returns
    -------
    pd.DataFrame(columns=['id', 'boiling', 'sku', 'kg', ...])
    """

    if isinstance(boiling_plan, pd.DataFrame):
        # already boiling plan
        return boiling_plan

    wb = cast_workbook(boiling_plan)

    cur_id = 0

    with code("Load boiling plan"):
        ws = None
        for key in ["План варок адыгейский", "План варок милкпроджект"]:
            if key in wb.sheetnames:
                ws = wb[key]
        if not ws:
            raise Exception("Не найдена вкладка для плана варок")

    values = []

    # collect header
    header = [ws.cell(1, i).value for i in range(1, 100) if ws.cell(1, i).value]

    for i in range(2, 200):
        if not ws.cell(i, 2).value:
            continue

        values.append([ws.cell(i, j).value for j in range(1, len(header) + 1)])

    df = pd.DataFrame(values, columns=header)
    df = df[
        [
            "Номер группы варок",
            "SKU",
            "КГ",
        ]
    ]
    df.columns = [
        "group_id",
        "sku",
        "kg",
    ]
    df = df[df["sku"] != "-"]
    df["group_id"] = df["group_id"].astype(int)

    # batch_id and boiling_id are the same with group_id
    df["batch_id"] = df["group_id"]
    df["boiling_id"] = df["group_id"]

    df["sku"] = df["sku"].apply(lambda sku: cast_model(MilkProjectSKU, sku))

    df["boiling"] = df["sku"].apply(lambda sku: delistify(sku.made_from_boilings, single=True))

    df["batch_type"] = "milk_project"
    df["absolute_batch_id"] = calc_absolute_batch_id(
        boiling_plan_df=df,
        first_batch_ids_by_type=first_batch_ids_by_type,
    )
    return df


def test():
    df = to_boiling_plan(
        str(get_repo_path() / "app/data/static/samples/by_department/milk_project/План по варкам милкпроджект 3.xlsx")
    )
    print(df.iloc[0])


if __name__ == "__main__":
    test()
