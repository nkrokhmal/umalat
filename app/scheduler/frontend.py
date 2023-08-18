from openpyxl.styles import Border, Font, NamedStyle, Side
from openpyxl.styles.borders import BORDER_THIN, Border, Side
from utils_ak.openpyxl.openpyxl_tools import *

from app.imports.runtime import *
from app.scheduler.time import *


def fill_grid(ws):
    utils.set_border_grid(ws, 1, 1, ws.max_column, ws.max_row, Side(border_style=BORDER_THIN))


def prepare_schedule_worksheet(ws_obj):
    ws = cast_worksheet(ws_obj)
    set_zoom(ws, 55)
    set_dimensions(ws, "column", range(1, 5), 21)
    set_dimensions(ws, "column", range(5, 288 * 2), 2.4)
    set_dimensions(ws, "row", range(1, 220), 25)
    return ws


def init_schedule_workbook(wb=None):
    if not wb:
        wb = utils.init_workbook(["Расписание"])

    if "Расписание" not in wb.sheetnames:
        wb.create_sheet("Расписание")

    prepare_schedule_worksheet(utils.cast_worksheet((wb, "Расписание")))
    return wb


def draw_schedule(schedule, style, O=None, fn=None, wb=None, debug=False, init=True):
    if init:
        wb = init_schedule_workbook(wb)
    logger.info("Finished init")
    ws = cast_worksheet((wb, "Расписание"))
    utils.set_active_sheet(wb, "Расписание")
    O = O or [0, 0]  # initial coordinates

    # update styles
    for b in schedule.iter():
        block_style = style.get(b.props["cls"])

        if block_style:
            block_style = {k: v(b) if callable(v) else v for k, v in block_style.items()}
            b.props.update(**block_style)

    schedule.props.update(index_width=4)

    for b in schedule.iter():
        logger.trace("Processing", b=b)
        if b.is_leaf() and b.props.get("visible", True):
            if b.size[0] == 0 or b.size[1] == 0:
                continue

            text = b.props.get("text", "")
            color = utils.cast_color(b.props.get("color", "white"))

            try:
                text = text.format(**b.props.all())
                text = text.replace("<", "{")
                text = text.replace(">", "}")
                text = eval(f"f{text!r}")

                # print(b.props['cls'], b.x, b.y)

                x1 = b.x[0]
                if "start_time" in b.props.all():
                    x1 -= cast_t(b.props["start_time"])  # shift of timeline
                x1 += b.props["index_width"]  # first index columns
                x1 += 1  # indexing starts with 1 in excel

                bold = b.props["bold"]
                font_size = b.props.get("font_size", 12)
                # print(b.props['cls'], x1, b.x[1], b.size[0], b.size[1])

                try:
                    utils.draw_merged_cell(
                        ws,
                        x1 + O[0],
                        b.x[1] + O[1],
                        b.size[0],
                        b.size[1],
                        text,
                        color,
                        bold=bold,
                        border=b.props.relative_props.get("border", {"border_style": "thin", "color": "000000"}),
                        text_rotation=b.props.get("text_rotation"),
                        font_size=font_size,
                        alignment="center",
                    )
                except AttributeError as e:
                    if "'MergedCell' object attribute 'value' is read-only" in str(e):
                        logger.exception("Block conflict during drawing")
                    if not debug:
                        raise
            except:
                logger.error("Failed to draw block", b=b, x=(x1, b.x[1]), size=b.size)
                logger.error("Relative props", props=b.props.relative_props)
                raise Exception(
                    "Ошибка в построении расписания, произошли накладки одних блоков на другие. Такого быть не должно, нужно обращатсья к разработчикам."
                )

    if fn:
        wb.save(fn)

    return wb


def draw_excel_frontend(frontend, style, O=None, open_file=False, fn=None, wb=None, init=True):
    wb = draw_schedule(frontend, style, O=O, wb=wb, init=init)

    if fn:
        utils.makedirs(fn)
        wb.save(fn)

        if open_file:
            utils.open_file_in_os(fn)

    return wb
