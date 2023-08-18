import os

import flask
import openpyxl

from utils_ak.os import makedirs

from app.scheduler.consolidated_schedule import run_consolidated
from app.scheduler.frontend import draw_excel_frontend
from app.scheduler.load_schedules import load_schedules
from app.scheduler.parsing import load_cells_df
from config import config


def run_consolidated_old(
    input_path, prefix="", output_path="outputs/", open_file=False, schedules=None, wb=None, date=None
):
    makedirs(output_path)

    if not schedules:
        schedules = load_schedules(input_path, prefix)

    cur_depth = 0

    if "mozzarella" in schedules:
        # init workbook with mozzarella
        if not wb:
            wb = openpyxl.load_workbook(
                filename=flask.current_app.config["TEMPLATE_SCHEDULE_PLAN_DEPARTMENT"],
                data_only=True,
            )

        from app.scheduler.mozzarella import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["mozzarella"])
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth
        draw_excel_frontend(frontend, STYLE, wb=wb)
    else:
        if input_path:
            # draw mozzarella
            if not wb:
                # init workbook with schedule file #todo maybe: better to crop mozzarella part and paste into empty workbook, so can be generalized for other departments
                wb = openpyxl.load_workbook(os.path.join(input_path, f"{prefix} Расписание моцарелла.xlsx"))

                for sheet_name in wb.sheetnames:
                    if sheet_name != "Расписание":
                        wb.remove(wb[sheet_name])

            df = load_cells_df(wb, "Расписание")
            cur_depth += df["y1"].max()

        if not wb:
            wb = openpyxl.load_workbook(
                filename=os.path.join(basedir, config.TEMPLATE_SCHEDULE_PLAN_DEPARTMENT),
                data_only=True,
            )

    if "ricotta" in schedules:
        from app.scheduler.ricotta import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["ricotta"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

        draw_excel_frontend(frontend, STYLE, wb=wb)

    if "mascarpone" in schedules:
        from app.scheduler.mascarpone import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["mascarpone"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

        draw_excel_frontend(frontend, STYLE, wb=wb)

    if "butter" in schedules:
        from app.scheduler.butter import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["butter"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

        draw_excel_frontend(frontend, STYLE, wb=wb)

    if "milk_project" in schedules:
        from app.scheduler.milk_project import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["milk_project"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

        draw_excel_frontend(frontend, STYLE, wb=wb)

    if "adygea" in schedules:
        from app.scheduler.adygea import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["adygea"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

        draw_excel_frontend(frontend, STYLE, wb=wb)

    if "contour_cleanings" in schedules:
        from app.scheduler.contour_cleanings import STYLE, wrap_frontend

        frontend = wrap_frontend(schedules["contour_cleanings"], date=date)
        depth = frontend.y[1]
        frontend.props.update(x=(frontend.x[0], frontend.x[1] + cur_depth))
        cur_depth += depth

    # todo maybe: copy-paste from submit_schedule
    with code("Dump frontend as excel file"):
        with code("Get filename"):
            output_fn = None
            if output_path:
                base_fn = f"Расписание общее.xlsx"
                if prefix:
                    base_fn = prefix + " " + base_fn
                output_fn = os.path.join(output_path, base_fn)

        draw_excel_frontend(frontend, open_file=open_file, wb=wb, fn=output_fn, style=STYLE)
    return wb


def test():
    run_consolidated(
        # "/Users/arsenijkadaner/Yandex.Disk.localized/master/code/git/2020.10-umalat/umalat/app/data/dynamic/2021-01-01/schedule_dict/",
        "/Users/marklidenberg/Yandex.Disk.localized/Загрузки/umalat/2021-07-16/approved",
        # prefix="2021-01-01",
        prefix="2021-07-16",
        open_file=True,
    )


if __name__ == "__main__":
    test()
