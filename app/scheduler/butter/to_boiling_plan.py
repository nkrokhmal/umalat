from typing import Union

import pandas as pd

from openpyxl import Workbook
from utils_ak.builtin import delistify
from utils_ak.openpyxl.openpyxl_tools import cast_workbook

from app.lessmore.utils.get_repo_path import get_repo_path
from app.models import ButterSKU, cast_model
from app.scheduler.boiling_plan_like import BoilingPlanLike
from app.scheduler.update_absolute_batch_id import update_absolute_batch_id


def to_boiling_plan(
    boiling_plan_source: BoilingPlanLike,
    first_batch_ids_by_type: dict = {"butter": 1},
) -> pd.DataFrame:
    """
    :param boiling_plan_source: str or openpyxl.Workbook
    :return: pd.DataFrame(columns=['id', 'boiling', 'sku', 'kg'])
    """

    if isinstance(boiling_plan_source, pd.DataFrame):
        # already a dataframe
        return boiling_plan_source

    # - Read boiling plan

    wb = cast_workbook(boiling_plan_source)

    cur_id = 0
    ws = wb["План варок"]

    values = []

    # collect header
    header = [ws.cell(1, i).value for i in range(1, 100) if ws.cell(1, i).value]

    for i in range(2, 200):
        if not ws.cell(i, 2).value:
            continue

        values.append([ws.cell(i, j).value for j in range(1, len(header) + 1)])

    df = pd.DataFrame(values, columns=header)
    df = df[["Номер группы варок", "SKU", "КГ"]]
    df.columns = ["group_id", "sku", "kg"]
    df = df[df["sku"] != "-"]

    df["group_id"] = df["group_id"].astype(int)

    # batch_id and boiling_id are the same with group_id
    df["batch_id"] = df["group_id"]
    df["boiling_id"] = df["group_id"]
    df["batch_type"] = "butter"
    df["sku"] = df["sku"].apply(lambda sku: cast_model(ButterSKU, sku))

    # - Saturate boiling plan

    df["boiling"] = df["sku"].apply(lambda sku: delistify(sku.made_from_boilings, single=True))
    df["start"] = None
    df["finish"] = None

    # - Update absolute batch id

    df = update_absolute_batch_id(boiling_plan_df=df, first_batch_ids_by_type=first_batch_ids_by_type)

    # - Return

    return df


def test():
    df = to_boiling_plan(
        str(get_repo_path() / "app/data/static/samples/by_department/butter/2023-09-03 План по варкам масло.xlsx")
    )
    print(df.iloc[0])


if __name__ == "__main__":
    test()
