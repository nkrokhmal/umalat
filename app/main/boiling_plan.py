from io import BytesIO
from flask import url_for, render_template, flash, request, make_response, current_app, request, send_from_directory
from pycel import ExcelCompiler
from werkzeug.utils import redirect
from . import main
import pandas as pd

from .errors import internal_error
from .. import db
from .forms import BoilingPlanForm, FileForm
from ..models_new import SKU, Boiling
from sqlalchemy import or_, and_
from flask_restplus import reqparse
import openpyxl
from app.generic_imports import *
import re
import numpy as np
from .. utils.generate_constructor import *


@main.route('/boiling_plan_full', methods=['POST', 'GET'])
def boiling_plan_full():
    # try:
        form = BoilingPlanForm()
        if request.method == 'POST' and form.validate_on_submit():
            file = request.files['input_file']
            file_path = os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename)
            if file:
                file.save(file_path)

            excel = ExcelCompiler(file_path)
            wb = openpyxl.load_workbook(filename=os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename),
                                        data_only=True)
            sheet_name = current_app.config['SHEET_NAMES']['schedule_plan']
            ws = wb[sheet_name]
            values = []
            for i in range(1, 200):
                if ws.cell(i, 4).value:
                    values.append([excel.evaluate("'{}'!{}".format(
                        sheet_name,
                        ws.cell(i, j).coordinate)) for j in range(4, 9)])

            df = pd.DataFrame(values[1:], columns=['sku', 'remainings - request', 'normative remainings', 'plan', 'extra_packing'])
            df_extra_packing = df[['sku', 'extra_packing']].copy()
            df = df.fillna(0)
            df = df[df['plan'] != 0]
            df = df[df['plan'].apply(lambda x: type(x) == int or type(x) == float or x.isnumeric())]
            path = '{}/{}.csv'.format(current_app.config['STATS_FOLDER'], os.path.splitext(file.filename)[0])
            df[['sku', 'plan']].to_csv(path, index=False)
            df = df[['sku', 'plan']]

            df['sku'] = df['sku'].apply(cast_sku)
            df = df.replace(to_replace='None', value=np.nan).dropna()
            df['boiling_id'] = df['sku'].apply(lambda x: x.made_from_boilings[0].id)
            df['sku_id'] = df['sku'].apply(lambda x: x.id)
            df['plan'] = df['plan'].apply(lambda x: abs(float(x)))
            df = df[['boiling_id', 'sku_id', 'plan']]
            df['sku'] = df['sku_id'].apply(cast_sku)
            full_plan = generate_constructor_df_v3(df)
            template_wb = openpyxl.load_workbook(file_path)
            new_file_name = draw_constructor_template(full_plan, file.filename, template_wb, df_extra_packing)
            if new_file_name != file_path:
                os.remove(file_path)
            return render_template('boiling_plan_full.html', form=form, file_name=new_file_name)
        file_name = None
        return render_template('boiling_plan_full.html', form=form, file_name=file_name)
    # except Exception as e:
    #     return internal_error(e)


