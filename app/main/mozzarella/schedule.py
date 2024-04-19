from app.imports.runtime import *  # isort: skip
import flask
import openpyxl

from utils_ak.openpyxl import write_metadata

from app.main import main
from app.main.errors import internal_error
from app.main.mozzarella.forms import ScheduleForm
from app.main.mozzarella.update_task_and_batches import update_task_and_batches
from app.main.validators import *
from app.scheduler import *
from app.scheduler.mozzarella import *
from app.scheduler.mozzarella.draw_frontend.draw_frontend import draw_frontend
from app.scheduler.mozzarella.to_boiling_plan.read_additional_packing import read_additional_packing
from app.scheduler.mozzarella.to_boiling_plan.to_boiling_plan import to_boiling_plan
from app.utils.batches.batch import *
from app.utils.features.openpyxl_wrapper import set_default_sheet
from app.utils.files.utils import create_if_not_exists, save_schedule, save_schedule_dict
from app.utils.mozzarella.additional_packing_draw import draw_additional_packing
from app.utils.mozzarella.boiling_plan_draw import draw_boiling_plan_merged
from app.utils.mozzarella.parse_schedule_json import *
from app.utils.mozzarella.schedule_task import MozzarellaScheduleTask


@main.route("/mozzarella_schedule", methods=["GET", "POST"])
@flask_login.login_required
def mozzarella_schedule():
    form = ScheduleForm(flask.request.form)
    if flask.request.method == "POST" and "submit" in flask.request.form:
        date = form.date.data
        add_full_boiling = form.add_full_boiling.data
        exact_melting_time_by_line = form.exact_melting_time_by_line.data
        rubber_beg_time = form.rubber_beg_time.data

        # validate time
        time_validator(form, form.water_beg_time)
        time_validator(form, form.salt_beg_time)

        file = flask.request.files["input_file"]
        data_dir = os.path.join(
            flask.current_app.config["DYNAMIC_DIR"],
            date.strftime("%Y-%m-%d"),
            flask.current_app.config["BOILING_PLAN_FOLDER"],
        )
        create_if_not_exists(data_dir)

        file_path = os.path.join(data_dir, file.filename)
        if file:
            file.save(file_path)
        wb = openpyxl.load_workbook(
            filename=os.path.join(data_dir, file.filename),
            data_only=True,
        )

        # Delete list "Расписание" if exists
        if "Расписание" in wb.sheetnames:
            wb.remove(wb["Расписание"])

        first_batch_ids = {"mozzarella": form.batch_number.data}
        boiling_plan_df = to_boiling_plan(wb, first_batch_ids_by_type=first_batch_ids)

        schedule_wb = openpyxl.load_workbook(filename=flask.current_app.config["TEMPLATE_SCHEDULE_PLAN"])

        try:
            output = draw_frontend(
                boiling_plan=file_path,
                start_times={
                    LineName.WATER: form.water_beg_time.data,
                    LineName.SALT: form.salt_beg_time.data,
                },
                exact_start_time_line_name=exact_melting_time_by_line,
                first_batch_ids_by_type=first_batch_ids,
                optimize_cleanings=add_full_boiling,
                date=date,
                workbook=schedule_wb,
            )

            schedule, schedule_wb = output["schedule"], output["workbook"]
        except Exception as e:
            return internal_error(e)

        additional_packing_df = read_additional_packing(wb)
        schedule_json = schedule.to_dict(
            props=[
                {"key": "x", "value": lambda b: list(b.props["x"])},
                {"key": "size", "value": lambda b: list(b.props["size"])},
                {"key": "cleaning_type", "cls": "cleaning"},
                {
                    "key": "boiling_group_df",
                    "cls": "boiling",
                },
            ]
        )
        cleanings = [x for x in schedule_json["children"][0]["children"] if x["cls"] == "cleaning"]
        schedule_df = prepare_schedule_json(schedule_json, cleanings)
        schedule_wb = draw_boiling_plan_merged(schedule_df, schedule_wb)
        schedule_wb = draw_additional_packing(schedule_wb, additional_packing_df)
        write_metadata(schedule_wb, json.dumps({"first_batch_ids": first_batch_ids, "date": str(date)}))

        schedule_task = update_task_and_batches(schedule_wb, boiling_plan_df=boiling_plan_df)

        schedule_wb = schedule_task.schedule_task_original(schedule_wb)
        schedule_wb = schedule_task.schedule_task_boilings(schedule_wb)

        set_default_sheet(schedule_wb)

        filename_schedule = "{} {}.xlsx".format(date.strftime("%Y-%m-%d"), "Расписание моцарелла")
        filename_schedule_pickle = "{} {}.pickle".format(date.strftime("%Y-%m-%d"), "Расписание моцарелла")

        save_schedule(schedule_wb, filename_schedule, date.strftime("%Y-%m-%d"))
        save_schedule_dict(schedule.to_dict(), filename_schedule_pickle, date.strftime("%Y-%m-%d"))

        return flask.render_template(
            "mozzarella/schedule.html",
            form=form,
            filename=filename_schedule,
            date=date.strftime("%Y-%m-%d"),
        )

    form.date.data = datetime.today() + timedelta(days=1)

    form.batch_number.data = (
        BatchNumber.last_batch_number(datetime.today() + timedelta(days=1), "Моцарельный цех", group="mozzarella") + 1
    )
    filename_schedule = None

    return flask.render_template("mozzarella/schedule.html", form=form, filename=filename_schedule)
