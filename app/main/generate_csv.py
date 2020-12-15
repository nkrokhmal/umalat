from io import BytesIO
from flask import url_for, render_template, flash, request, make_response, current_app, request
from werkzeug.utils import redirect
from . import main
import pandas as pd
from .. import db
from .forms import StatisticForm
from ..models import SKU, Boiling
from sqlalchemy import or_, and_
from flask_restplus import reqparse
import openpyxl
from app.interactive_imports import *
import re


@main.route('/generate_stats_new', methods=['POST', 'GET'])
def generate_stats_new():
    pass


# todo: add plan
# todo: file naming
@main.route('/generate_stats_old', methods=['POST', 'GET'])
def generate_stats_old():
    form = StatisticForm()
    if request.method == 'POST' and form.validate_on_submit():
        file = request.files['input_file']
        file_name = file.filename
        file_data = BytesIO(file.read())
        wb = openpyxl.load_workbook(filename=file_data)
        ws = wb['планирование суточное']
        values = []
        for i in range(1, 200):
            if ws.cell(i, 3).value:
                values.append([ws.cell(i, j).value for j in range(3, 7)])

        df = pd.DataFrame(values[1:], columns=['sku', 'remainings - request', 'normative remainings', 'plan'])
        df = df.fillna(0)
        df = df[df['plan'] != 0]
        path = '{}/{}.csv'.format('app/data/stats', os.path.splitext(file_name)[0])
        df[['sku', 'plan']].to_csv(path, index=False)
        return render_template('stats_old.html', form=form)
    return render_template('stats_old.html', form=form)
