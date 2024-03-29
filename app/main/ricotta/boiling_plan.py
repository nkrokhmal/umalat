import os

from datetime import datetime, timedelta
from pathlib import Path

import flask
import flask_login
import openpyxl

from app.globals import db
from app.main import main
from app.main.ricotta.forms import BoilingPlanForm
from app.models.basic import SKU
from app.models.ricotta import RicottaBoiling, RicottaSKU
from app.scheduler.mozzarella.to_boiling_plan.to_boiling_plan import to_boiling_plan
from app.utils.files.utils import move_boiling_file, save_request
from app.utils.parse_remainings import get_skus, group_skus, parse_file_path, parse_sheet
from app.utils.ricotta.boiling_plan_create import boiling_plan_create
from app.utils.ricotta.boiling_plan_draw import draw_boiling_plan
from app.utils.sku_plan import SkuPlanClient


@main.route("/ricotta_boiling_plan", methods=["POST", "GET"])
@flask_login.login_required
def ricotta_boiling_plan():
    form = BoilingPlanForm(flask.request.form)
    if flask.request.method == "POST" and "submit" in flask.request.form:
        date: datetime = form.date.data
        add_auto_boilings: bool = form.add_auto_boilings.data

        skus = db.session.query(RicottaSKU).all()
        total_skus = db.session.query(SKU).all()
        boilings = db.session.query(RicottaBoiling).all()
        total_volume = request_kg = 0

        file = flask.request.files["input_file"]
        filename_ext = file.filename.split(".")[-1]

        upload_path = Path(flask.current_app.config["UPLOAD_TMP_FOLDER"]) / file.filename

        if file:
            file.save(upload_path)

        if filename_ext == "csv":
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(filename=upload_path, data_only=True)
            if "Вода" in wb.sheetnames and "Соль" in wb.sheetnames:
                boiling_plan_df = to_boiling_plan(wb)
                boiling_plan_df["configuration"] = boiling_plan_df["configuration"].apply(lambda x: int(x))
                total_volume = int(boiling_plan_df.groupby("group_id").first()["configuration"].sum() * 0.81)

                if add_auto_boilings:
                    request_kg = total_volume

        skus_req, remainings_df = parse_file_path(str(upload_path))

        skus_req = get_skus(skus_req, skus, total_skus)
        skus_grouped = group_skus(skus_req, boilings)
        sku_plan_client = SkuPlanClient(
            date=date,
            remainings=remainings_df,
            skus_grouped=skus_grouped,
            template_path=flask.current_app.config["TEMPLATE_RICOTTA_BOILING_PLAN"],
        )
        sku_plan_client.fill_remainings_list()
        sku_plan_client.fill_ricotta_sku_plan()

        excel_compiler, wb, wb_data_only, filename, filepath = move_boiling_file(
            sku_plan_client.date,
            sku_plan_client.filepath,
            sku_plan_client.filename,
            "рикотта",
        )
        sheet_name = flask.current_app.config["SHEET_NAMES"]["schedule_plan"]
        ws = wb_data_only[sheet_name]
        df, _ = parse_sheet(ws, sheet_name, excel_compiler, RicottaSKU)
        df_plan = boiling_plan_create(df, request_kg)

        wb = draw_boiling_plan(df_plan, wb, total_volume)
        save_request(data=wb, filename=filename, date=sku_plan_client.date)
        os.remove(upload_path)
        return flask.render_template(
            "ricotta/boiling_plan.html", form=form, filename=filename, date=sku_plan_client.date
        )
    form.date.data = datetime.today() + timedelta(days=1)
    return flask.render_template("ricotta/boiling_plan.html", form=form, filename=None)
