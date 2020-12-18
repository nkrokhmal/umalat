from io import BytesIO
from flask import url_for, render_template, flash, request, make_response, current_app, request
from werkzeug.utils import redirect
from . import main
import pandas as pd
from .. import db
from .forms import StatisticForm
from ..models import SKU, Boiling
from sqlalchemy import or_, and_
from flask_restplus import reqparse
import openpyxl
from app.interactive_imports import *
import re
from pycel import ExcelCompiler


# todo: add naming
# todo: add plan
@main.route('/sku_plan_stats_new', methods=['POST', 'GET'])
def sku_plan_stats_new():
    form = StatisticForm()
    if request.method == 'POST' and form.validate_on_submit():
        file = request.files['input_file']
        file_path = os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename)
        if file:
            file.save(file_path)

        excel = ExcelCompiler(file_path)
        wb = openpyxl.load_workbook(filename=os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename),
                                    data_only=True)
        # todo: use config
        sheet_name = 'планирование суточное'
        ws = wb[sheet_name]
        values = []
        for i in range(1, 200):
            if ws.cell(i, 4).value:
                values.append([excel.evaluate("'{}'!{}".format(
                    sheet_name,
                    ws.cell(i, j).coordinate)) for j in range(4, 8)])

        df = pd.DataFrame(values[1:], columns=['sku', 'remainings - request', 'normative remainings', 'plan'])
        df = df.fillna(0)
        df = df[df['plan'] != 0]
        df = df[df['plan'].apply(lambda x: type(x) == int or type(x) == float or x.isnumeric())]
        path = '{}/{}.csv'.format(current_app.config['STATS_FOLDER'], os.path.splitext(file.filename)[0])
        df[['sku', 'plan']].to_csv(path, index=False)
        os.remove(file_path)
        return render_template('sku_plan_stats_new.html', form=form, filename='{}.csv'.format(os.path.splitext(file.filename)[0]))
    filename = None
    return render_template('sku_plan_stats_new.html', form=form, filename=filename)


# todo: add plan
# todo: file naming
@main.route('/sku_plan_stats_old', methods=['POST', 'GET'])
def sku_plan_stats_old():
    form = StatisticForm()
    if request.method == 'POST' and form.validate_on_submit():
        file = request.files['input_file']
        file_path = os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename)
        if file:
            file.save(file_path)

        excel = ExcelCompiler(file_path)
        wb = openpyxl.load_workbook(filename=os.path.join(current_app.config['UPLOAD_TMP_FOLDER'], file.filename),
                                    data_only=True)
        sheet_name = 'планирование суточное'
        ws = wb[sheet_name]
        values = []
        for i in range(1, 200):
            if ws.cell(i, 3).value:
                values.append([excel.evaluate("'{}'!{}".format(
                    sheet_name,
                    ws.cell(i, j).coordinate)) for j in range(3, 7)])

        df = pd.DataFrame(values[1:], columns=['sku', 'remainings - request', 'normative remainings', 'plan'])
        df = df.fillna(0)
        df = df[df['plan'] != 0]
        df = df[df['plan'].apply(lambda x: type(x) == int or type(x) == float or x.isnumeric())]
        path = '{}/{}.csv'.format(current_app.config['STATS_FOLDER'], os.path.splitext(file.filename)[0])
        df[['sku', 'plan']].to_csv(path, index=False)
        os.remove(file_path)
        return render_template('sku_plan_stats_old.html', form=form, filename='{}.csv'.format(os.path.splitext(file.filename)[0]))
    filename = None
    return render_template('sku_plan_stats_old.html', form=form, filename=filename)
