from app.imports.runtime import *

from app.models import *
from app.schedule_maker.departments.ricotta.boiling_plan.saturate import (
    saturate_boiling_plan,
)

from openpyxl.utils.cell import column_index_from_string


INDEX_COLUMN = column_index_from_string("A")
# todo: check
KG_COLUMN = column_index_from_string("F") - 1
TANKS_COLUMN = column_index_from_string("D")
OUTPUT_COLUMN = column_index_from_string("C")
NAME_COLUMN = column_index_from_string("E")
TANKS_TOTAL_COLUMN = column_index_from_string("Q")
FIRST_TANK_COLUMN = column_index_from_string("S")


class BoilingParser:
    def __init__(
        self,
        boilings,
        number_of_tanks,
        kg_per_tank,
        number_of_tanks_in_boiling,
        number_of_boiling=1,
        first_tank_number="",
    ):
        self.boilings = boilings
        self.boilings_copy = boilings
        self.number_of_tanks = number_of_tanks
        self.kg_per_tank = kg_per_tank
        self.number_of_tanks_in_boiling = number_of_tanks_in_boiling
        self.index = 0
        self.number_of_boiling = number_of_boiling
        self.first_tank_number = first_tank_number

    def add_tank(self):
        add_boiling = []
        if self.index > len(self.boilings) - 1:
            return add_boiling
        cur_boiling = self.boilings[self.index]
        while True:
            if self.index == len(self.boilings) - 1:
                if cur_boiling[KG_COLUMN] == 0:
                    break
            if cur_boiling[KG_COLUMN] >= self.kg_per_tank - sum(
                [x[KG_COLUMN] for x in add_boiling]
            ):
                boiling = deepcopy(cur_boiling)
                boiling[KG_COLUMN] = self.kg_per_tank - sum(
                    [x[KG_COLUMN] for x in add_boiling]
                )
                self.boilings[self.index][KG_COLUMN] -= self.kg_per_tank - sum(
                    [x[KG_COLUMN] for x in add_boiling]
                )
                add_boiling.append(boiling)
                break
            else:
                add_boiling.append(deepcopy(cur_boiling))
                self.boilings[self.index][KG_COLUMN] = 0
                self.index += 1
                if self.index > len(self.boilings) - 1:
                    break
                cur_boiling = self.boilings[self.index]
        return add_boiling

    def parse(self):
        result = []
        current_number_of_tanks = 0
        is_first = True
        while current_number_of_tanks < self.number_of_tanks:
            tanks = min(
                self.number_of_tanks_in_boiling,
                self.number_of_tanks - current_number_of_tanks,
            )
            boiling_result = []
            for _ in range(tanks):
                boiling_result += self.add_tank()
            if boiling_result:
                boiling_result = pd.DataFrame(boiling_result)
                groupby_columns = list(range(KG_COLUMN))
                groupby_columns.append(FIRST_TANK_COLUMN - 1)
                boiling_result[FIRST_TANK_COLUMN - 1] = (
                    self.first_tank_number if is_first else ""
                )

                boiling_result = (
                    pd.DataFrame(boiling_result)
                    .groupby(groupby_columns)
                    .agg({KG_COLUMN: "sum"})
                    .reset_index()
                )

                boiling_result[INDEX_COLUMN - 1] = self.number_of_boiling
                boiling_result[TANKS_COLUMN - 1] = tanks

                result += boiling_result.values.tolist()
                self.number_of_boiling += 1
                current_number_of_tanks += tanks
                is_first = False
            else:
                raise AssertionError(
                    "В одной из варок неправильно указано число кг, не хватает на указанное число варок"
                )
        return result


def read_boiling_plan(wb_obj, saturate=True):
    """
    :param wb_obj: str or openpyxl.Workbook
    :return: pd.DataFrame(columns=['id', 'boiling', 'sku', 'kg'])
    """
    wb = utils.cast_workbook(wb_obj)

    ws_name = "План варок"
    ws = wb[ws_name]
    values = []

    # collect header
    length = 20
    current_boiling = []
    boiling_index_number = 1

    for i in range(2, 200):
        if not ws.cell(i, 2).value:
            continue
        if ws.cell(i, NAME_COLUMN).value == "-":
            if isinstance(ws.cell(i, TANKS_TOTAL_COLUMN).value, int):
                first_tank_number = ws.cell(i, FIRST_TANK_COLUMN).value or ""
                parser = BoilingParser(
                    boilings=current_boiling,
                    number_of_tanks=ws.cell(i, TANKS_TOTAL_COLUMN).value,
                    number_of_tanks_in_boiling=number_of_tanks_in_boiling,
                    kg_per_tank=kg_per_tank,
                    number_of_boiling=boiling_index_number,
                    first_tank_number=first_tank_number,
                )
                values += parser.parse()
                boiling_index_number = parser.number_of_boiling
                current_boiling = []
        else:
            current_boiling.append([ws.cell(i, j).value for j in range(1, length)])
            number_of_tanks_in_boiling = ws.cell(i, TANKS_COLUMN).value
            kg_per_tank = float(
                ws.cell(i, OUTPUT_COLUMN).value / number_of_tanks_in_boiling
            )

    df = pd.DataFrame(
        values,
        columns=[
            "boiling_id",
            "boiling_type",
            "output",
            "tanks",
            "sku",
            "first_tank",
            "kg",
        ],
    )
    df["sku"] = df["sku"].apply(lambda sku: cast_model(RicottaSKU, sku))
    df["boiling"] = df["sku"].apply(lambda sku: sku.made_from_boilings[0])
    df["full_tank_number"] = df["sku"].apply(
        lambda sku: sku.made_from_boilings[0].number_of_tanks
    )
    df["boiling_id"] = df["boiling_id"].astype(int)
    df["kg"] = df["kg"].apply(lambda x: int(x))

    for idx, grp in df.groupby("boiling_id"):
        assert (
            len(grp["boiling"].unique()) == 1
        ), "В одной варке должен быть только один тип варки"

        assert (
            len(grp["output"].unique()) == 1
        ), "В одной варке должны совпадать выходы с варки"

        # fix number of kilograms
        if (
            abs(
                grp["kg"].sum()
                - grp.iloc[0]["output"]
                * grp.iloc[0]["tanks"]
                / grp.iloc[0]["full_tank_number"]
            )
            / grp.iloc[0]["output"]
            > 0.05
        ):
            raise AssertionError(
                "Одна из групп варок имеет неверное количество килограмм."
            )
        else:
            if (
                abs(
                    grp["kg"].sum()
                    - grp.iloc[0]["output"]
                    * grp.iloc[0]["tanks"]
                    / grp.iloc[0]["full_tank_number"]
                )
                > 1e-5
            ):
                df.loc[grp.index, "kg"] *= (
                    grp.iloc[0]["output"] / grp["kg"].sum()
                )  # scale to total_volume
            else:
                # all fine
                pass
    if saturate:
        df = saturate_boiling_plan(df)
    return df
