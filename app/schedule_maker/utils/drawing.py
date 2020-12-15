import os
import openpyxl as opx
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter


from utils_ak.interactive_imports import *


from app.schedule_maker.utils.color import cast_color
from app.schedule_maker.utils.time import cast_t, cast_time
from app.schedule_maker.utils.interval import cast_interval, calc_interval_length
from app.schedule_maker.blocks import make_template

def set_border(sheet, x, y, w, h, border):
    rows = sheet['{}{}'.format(get_column_letter(x), y):'{}{}'.format(get_column_letter(x + w - 1), y + h - 1)]

    for row in rows:
        row[0].border = Border(left=border, top=row[0].border.top, bottom=row[0].border.bottom, right=row[0].border.right)
        row[-1].border = Border(left=row[-1].border.left, top=row[-1].border.top, bottom=row[-1].border.bottom, right=border)
    for c in rows[0]:
        c.border = Border(left=c.border.left, top=border, bottom=c.border.bottom, right=c.border.right)
    for c in rows[-1]:
        c.border = Border(left=c.border.left, top=c.border.top, bottom=border, right=c.border.right)


def draw_cell(sheet, x, y, text, color=None, font_size=None, text_rotation=None, alignment=None):
    cell = sheet.cell(row=y, column=x)
    cell.font = Font(size=font_size)
    if alignment == 'center':
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=text_rotation)
    cell.value = text

    if color:
        cell.fill = PatternFill("solid", fgColor=color[1:])
    return cell


def draw_block(sheet, x, y, w, h, text, color=None, border=None, text_rotation=None, font_size=None, alignment=None):
    color = color or cast_color('white')
    sheet.merge_cells(start_row=y, start_column=x, end_row=y + h - 1, end_column=x + w - 1)
    merged_cell = sheet.cell(row=y, column=x)
    merged_cell.font = Font(size=font_size)
    if alignment == 'center':
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=text_rotation)
    merged_cell.value = text
    if color:
        merged_cell.fill = PatternFill("solid", fgColor=color[1:])

    if border is not None:
        if isinstance(border, dict):
            border = Side(**border)
        elif isinstance(border, Side):
            pass
        else:
            raise Exception('Unknown border type')

        set_border(sheet, x, y, w, h, border)


def draw(sheet, block):
    for b in block.iter():
        if not b.children:
            text = b.props.get('text', '')
            color = cast_color(b.props.get('color', 'white'))

            if b.props.get('visible') == False:
                continue
            try:
                text = text.format(**b.props.get_all_props())
                text = text.replace('<', '{')
                text = text.replace('>', '}')
                text = eval(f'f{text!r}')

                beg = b.props['t']
                beg -= cast_t(b.props['beg_time'])  # shift of timeline
                beg += b.props['index_width']  # first index columns
                beg += 1  # indexing starts with 1 in excel

                draw_block(sheet, beg, b.props['y'], b.size, b.props.get('h', 1), text, color, border={'border_style': 'thin', 'color': '000000'}, text_rotation=b.props.get('text_rotation'))
            except:
                print(b)
                print(b.props.relative_props, b.interval)
                raise


def init_empty_sheet():
    work_book = opx.Workbook()
    sheet = work_book.worksheets[0]
    return work_book, sheet


def init_sheets(*args):
    workbook = opx.Workbook()
    for i, arg in enumerate(args):
        workbook.create_sheet(arg, i)
    return workbook


def init_sheet():
    work_book, sheet = init_empty_sheet()
    for i in range(4):
        sheet.column_dimensions[get_column_letter(i + 1)].width = 12
    for i in range(4, 288 * 2):
        sheet.column_dimensions[get_column_letter(i + 1)].width = 1.2
    return work_book, sheet


def init_template_sheet(template_fn=None):
    def _init_sheet():
        if template_fn:
            work_book = opx.load_workbook(template_fn)
            return work_book, work_book.worksheets[0]
        else:
            wb, sheet = init_sheet()
            draw(sheet, make_template())
            return wb, sheet
    return _init_sheet


def draw_row(sheet, y, values, color=None, **kwargs):
    for i, v in enumerate(values, 1):
        draw_cell(sheet, i, y, text=v, color=color, **kwargs)


def draw_schedule(root, style, fn=None, init_sheet_func=init_sheet):
    # update styles
    for b in root.iter():
        block_style = style.get(b.props['class'])

        if block_style:
            block_style = {k: v(b) if callable(v) else v for k, v in block_style.items()}
            b.props.update(block_style)

    root.props.update({'index_width': 4})

    work_book, sheet = init_sheet_func()

    draw(sheet, root)
    if fn:
        work_book.save(fn)
    return work_book


def draw_print(block, visible_only=True):
    res = ''
    for b in block.iter():
        if calc_interval_length(b.interval) != 0:
            if visible_only and b.props['visible'] is False:
                continue
            res += ' ' * b.beg + '=' * int(calc_interval_length(b.interval)) + f' {b.props["class"]} {b.props["y"]}:{b.interval}'
            res += '\n'
    return res


if __name__ == '__main__':
    print(init_template_sheet(None)())