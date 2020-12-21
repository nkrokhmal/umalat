import openpyxl
import pandas as pd

from app.schedule_maker.utils.block import *
from app.schedule_maker.utils.drawing import *
from app.schedule_maker.style import *
from app.schedule_maker.models import *
from app.schedule_maker.algo import *


def load_boiling_plan(wb_obj):
    if isinstance(wb_obj, str):
        wb = openpyxl.load_workbook(filename=wb_obj, data_only=True)
    elif isinstance(wb_obj, openpyxl.Workbook):
        wb = wb_obj
    else:
        raise Exception('Unknown workbook format')

    dfs = []

    for ws_name in ['Вода', 'Соль']:
        ws = wb[ws_name]
        values = []
        for i in range(2, 200):
            if not ws.cell(i, 1).value:
                continue
            values.append([ws.cell(i, j).value for j in [1, 2, 6, 7]])
        df = pd.DataFrame(values, columns=['boiling', 'id', 'sku', 'kg']) # first value is header
        df = df[df['boiling'] != '-']
        df = df[~df['kg'].isnull()]
        df = df[['id', 'boiling', 'sku', 'kg']] # reorder
        if dfs:
            df['id'] = df['id'] + dfs[-1].iloc[-1]['id']
        dfs.append(df)

    df = pd.concat(dfs)
    boiling_plan_df = df
    boiling_plan_df['sku'] = boiling_plan_df['sku'].apply(cast_sku)
    boiling_plan_df['boiling'] = boiling_plan_df['sku'].apply(lambda sku: sku.boilings[0])
    return boiling_plan_df


def add_push_with_props(parent, block):
    return add_push(parent, block, new_props=block.props.get_all_props())


def generate_schedule_frontend_block(root):
    maker = BlockMaker(default_push_func=dummy_push)
    make = maker.make

    with make('schedule'):
        make(h=1, visible=False, push_func=dummy_push_y)

        with make('header', h=1, index_width=3, push_func=dummy_push_y):
            make(h=1, size=1, text='График наливов', push_func=add_push)
            for i in range(288):
                cur_time = cast_time(i + cast_t('01:00'))
                if cur_time[-2:] == '00':
                    make(t=1 + i, size=1, h=1, text=str(int(cur_time[:2])), color=(218, 150, 148), text_rotation=90, push_func=add_push)
                else:
                    make(t=1 + i, size=1, h=1, text=cur_time[-2:], color=(204, 255, 255), text_rotation=90, push_func=add_push)

        for i in range(2):
            with make(f'cheese_maker_{i}', h=2, push_func=dummy_push_y):
                for block in root['boiling']:
                    if block.props['pouring_line'] == str(i):
                        make(block['pouring'], push_func=add_push_with_props)
            make(h=1, visible=False, push_func=dummy_push_y)

        with make('cleanings', h=2, push_func=dummy_push_y):
            for block in root['cleaning']:
                make(block, push_func=add_push_with_props)
        make(h=1, visible=False, push_func=dummy_push_y)

        for i in range(2, 4):
            with make(f'cheese_maker_{i}', h=2, push_func=dummy_push_y):
                for block in root['boiling']:
                    if block.props['pouring_line'] == str(i):
                        make(block['pouring'], push_func=add_push_with_props)
            make(h=1, visible=False, push_func=dummy_push_y)

        with make('header', h=1, index_width=4, push_func=dummy_push_y):
            for i in range(288):
                cur_time = cast_time(i + cast_t('07:00'))
                if cur_time[-2:] == '00':
                    make(t=i, size=1, h=1, text=str(int(cur_time[:2])), color=(218, 150, 148), text_rotation=90, push_func=add_push)
                else:
                    make(t=i, size=1, h=1, text=cur_time[-2:], color=(204, 255, 255), text_rotation=90, push_func=add_push)

        with make('water_melting', h=4, push_func=dummy_push_y):
            for block in root['boiling']:
                if block.props['boiling_type'] == 'water':
                    make(block['melting_and_packing']['melting'], push_func=add_push_with_props)

        with make('water_packing', h=3, push_func=dummy_push_y):
            for block in root['boiling']:
                if block.props['boiling_type'] == 'water':
                    make(block['melting_and_packing']['packing_and_preconfiguration'], push_func=add_push_with_props)

        # todo: make dynamic lines
        n_lines = 5
        melting_lines = [make(f'salt_melting_{i}', h=3, push_func=dummy_push_y).block for i in range(n_lines)]

        # todo: hardcode, add empty elements for drawing not to draw melting_line itself
        for b in melting_lines:
            add_push(b, Block('block', visible=False))

        for i, block in enumerate([b for b in root['boiling'] if b.props['boiling_type'] == 'salt']):
            add_push_with_props(melting_lines[i % n_lines], block['melting_and_packing']['melting'])

        with make('salt_packing', h=3, push_func=dummy_push_y):
            for block in root['boiling']:
                if block.props['boiling_type'] == 'salt':
                    make(block['melting_and_packing']['packing_and_preconfiguration'], push_func=add_push_with_props)
    return maker.root


def create_excel_frontend(boiling_plan_df):
    schedule = make_schedule(boiling_plan_df)
    frontend_block = generate_schedule_frontend_block(schedule)
    style = load_style()
    init_sheet_func = init_sheet
    wb = draw_schedule(frontend_block, style, init_sheet_func=init_sheet_func)
    return wb
