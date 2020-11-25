import openpyxl as opx
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter


from src.color import cast_color
from src.time import cast_t, cast_time
from src.interval import cast_interval, calc_interval_length


def set_border(sheet, x, y, w, h, border):
    rows = sheet['{}{}'.format(get_column_letter(x), y):'{}{}'.format(get_column_letter(x + w - 1), y + h - 1)]

    for row in rows:
        row[0].border = Border(left=border, top=row[0].border.top, bottom=row[0].border.bottom, right=row[0].border.right)
        row[-1].border = Border(left=row[-1].border.left, top=row[-1].border.top, bottom=row[-1].border.bottom, right=border)
    for c in rows[0]:
        c.border = Border(left=c.border.left, top=border, bottom=c.border.bottom, right=c.border.right)
    for c in rows[-1]:
        c.border = Border(left=c.border.left, top=c.border.top, bottom=border, right=c.border.right)


def draw_block(sheet, x, y, w, h, text, colour, border=None):
    if not colour:
        colour = cast_color('white')  # default white colour
    sheet.merge_cells(start_row=y, start_column=x, end_row=y + h - 1, end_column=x + w - 1)
    merged_cell = sheet.cell(row=y, column=x)
    merged_cell.value = text
    merged_cell.alignment = Alignment(horizontal='center')
    merged_cell.fill = PatternFill("solid", fgColor=colour[1:])

    if border is not None:
        if isinstance(border, dict):
            border = Side(**border)
        elif isinstance(border, Side):
            pass
        else:
            raise Exception('Unknown border type')

        set_border(sheet, x, y, w, h, border)


def draw(sheet, block):
    for b in block.iter():
        if not b.children:
            text = b.abs_props.get('text', '')
            color = cast_color(b.abs_props.get('color', 'white'))

            if b.abs_props.get('visible') == False:
                continue

            text = text.format(**b.abs_props)
            text = text.replace('<', '{')
            text = text.replace('>', '}')
            text = eval(f'f{text!r}')

            beg = b.abs_props['t']
            beg -= cast_t(b.abs_props['beg_time'])  # shift of timeline
            beg += b.abs_props['index_width']  # first index columns
            beg += 1  # indexing starts with 1 in excel

            print(b.abs_props['class'], b.abs_props['y'], cast_interval(beg, beg + b.size))
            draw_block(sheet, beg, b.abs_props['y'], b.size, 1, text, color, border={'border_style': 'thin', 'color': '000000'})


def init_sheet():
    work_book = opx.Workbook()
    sheet = work_book.worksheets[0]
    for i in range(288):
        sheet.column_dimensions[get_column_letter(i + 1)].width = 1.2
    return work_book, sheet


def init_template_sheet():
    def move_sheet(wb, from_loc=None, to_loc=None):
        sheets = wb._sheets

        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1

        # if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0

        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)

    work_book = opx.load_workbook(r'2020.11.18 schedule_template.xlsx')

    # delete all but last - original
    for sheet in work_book.worksheets[:-1]:
        work_book.remove_sheet(sheet)

    # copy sheet to work with
    sheet = work_book.copy_worksheet(work_book.worksheets[0])
    # put it at the beginning
    move_sheet(work_book, 1, 0)

    return work_book, sheet


def draw_schedule(root, styles, stylings, fn='output.xlsx', init_sheet_func=init_sheet):
    # update styles
    for b in root.iter():
        block_style = styles.get(b.rel_props['class'])
        if block_style:
            b.rel_props.update(block_style)

        block_styling = stylings.get(b.rel_props['class'])
        if block_styling:
            b.rel_props.update(block_styling(b))

    work_book, sheet = init_sheet_func()
    root.rel_props['index_width'] = 4
    draw(sheet, root)
    work_book.save(fn)


def draw_print(block):
    res = ''
    for b in block.iter():
        if calc_interval_length(b.interval) != 0:
            res += ' ' * int(b.abs_props['t']) + '=' * int(calc_interval_length(b.interval)) + f' {b.rel_props["class"]} {b.interval}'
            res += '\n'
    return res